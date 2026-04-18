#!/usr/bin/env python3
"""
Montana Segurança Privada — Apuração Mensal PIS/COFINS
Regime: Lucro Real Anual — Cumulativo (PIS 0,65% + COFINS 3,00%)
Base: Regime de Caixa a partir de jan/2026
Regra de transição: recebimentos de NFs emitidas antes de 2026 já foram
  tributados por competência — são EXCLUÍDOS da base de cálculo.

Uso:
  python scripts/apuracao_piscofins_seguranca_mensal.py --ano-mes=2026-03

Saída:
  output/Apuracao_PISCOFINS_Seguranca_202603.xlsx
"""

import argparse
import calendar
import sqlite3
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl não encontrado.")
    print("      Instale com: pip install openpyxl")
    sys.exit(1)

# ─── Caminhos ─────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent
DB_PATH  = BASE_DIR / 'data' / 'seguranca' / 'montana.db'
OUT_DIR  = BASE_DIR / 'output'

# ─── Alíquotas — regime cumulativo ────────────────────────────────
ALIQ_PIS    = 0.0065   # 0,65%
ALIQ_COFINS = 0.030    # 3,00%
DARF_PIS    = '8109'
DARF_COFINS = '2172'

# Início do regime de caixa (NFs emitidas antes disso = excluídas)
INICIO_CAIXA = '2026-01-01'

# Palavras-chave para créditos que NÃO tributam receita operacional
NAO_TRIBUTA_KW = [
    'rende facil', 'rende fácil', 'rende-facil',
    'ted proprio', 'ted próprio',
    'transf interna', 'transferencia interna', 'transferência interna',
    'aplicação', 'aplicacao',
    'resgate',
    'brb invest',
    'poupança', 'poupanca',
    'montana assessoria', 'montana serviços', 'montana servicos',
    'estorno ted', 'dev ted',
]


# ─── Helpers ──────────────────────────────────────────────────────
def brl(v: float) -> str:
    v = v or 0.0
    s = f"{v:,.2f}"                        # 1,234.56
    s = s.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"R$ {s}"


def calc_vencimento(ano: int, mes: int) -> date:
    """Vencimento DARF: dia 25 do mês seguinte (ou próximo dia útil)."""
    mes_seg = mes % 12 + 1
    ano_seg = ano + (1 if mes == 12 else 0)
    d = date(ano_seg, mes_seg, 25)
    while d.weekday() >= 5:   # sábado=5, domingo=6
        d += timedelta(days=1)
    return d


def is_nao_tributa(historico: str) -> bool:
    h = (historico or '').lower()
    return any(k in h for k in NAO_TRIBUTA_KW)


# ─── Apuração ─────────────────────────────────────────────────────
def apurar(ano_mes: str) -> dict:
    ano, mes = map(int, ano_mes.split('-'))
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    date_from = f"{ano}-{mes:02d}-01"
    date_to   = f"{ano}-{mes:02d}-{ultimo_dia:02d}"

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Todos os créditos bancários do mês, com NF vinculada (se houver)
    cur.execute("""
        SELECT
            e.id, e.data_iso, e.historico, e.credito,
            e.pagador_identificado, e.pagador_cnpj,
            e.status_conciliacao, e.contrato_vinculado,
            nf.id            AS nf_id,
            nf.numero        AS nf_numero,
            nf.data_emissao  AS data_emissao,
            nf.competencia   AS nf_competencia,
            nf.tomador,
            nf.valor_bruto   AS nf_valor_bruto,
            nf.valor_liquido AS nf_valor_liquido
        FROM extratos e
        LEFT JOIN notas_fiscais nf ON nf.extrato_id = e.id
        WHERE e.data_iso >= ? AND e.data_iso <= ?
          AND e.credito > 0
        ORDER BY e.data_iso
    """, (date_from, date_to))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    tributaveis = []
    excluidos   = []
    nao_tributa = []
    pendentes   = []

    for r in rows:
        # Período da NF: usa data_emissao (AAAA-MM-DD); fallback: competencia (AAAA-MM)
        nf_periodo = (r['data_emissao'] or r['nf_competencia'] or '').strip()

        if r['nf_id']:
            # NF vinculada: tributável só se emitida em jan/2026+
            if nf_periodo >= INICIO_CAIXA:
                tributaveis.append(r)
            else:
                excluidos.append(r)
        elif is_nao_tributa(r['historico']):
            nao_tributa.append(r)
        else:
            pendentes.append(r)

    base = sum(r['credito'] or 0 for r in tributaveis)
    pis_val    = round(base * ALIQ_PIS,    2)
    cofins_val = round(base * ALIQ_COFINS, 2)
    total_darf = round(pis_val + cofins_val, 2)
    vcto       = calc_vencimento(ano, mes)

    return {
        'ano_mes': ano_mes, 'ano': ano, 'mes': mes,
        'date_from': date_from, 'date_to': date_to,
        'base_tributavel': base,
        'pis': pis_val, 'cofins': cofins_val, 'total_darf': total_darf,
        'vencimento': vcto.strftime('%d/%m/%Y'),
        'tributaveis': tributaveis,
        'excluidos':   excluidos,
        'nao_tributa': nao_tributa,
        'pendentes':   pendentes,
    }


# ─── Geração do Excel ─────────────────────────────────────────────
def gerar_excel(d: dict) -> Path:
    OUT_DIR.mkdir(exist_ok=True)
    slug     = d['ano_mes'].replace('-', '')
    out_path = OUT_DIR / f"Apuracao_PISCOFINS_Seguranca_{slug}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Estilos ───────────────────────────────────────────────────
    HDR_FILL  = PatternFill('solid', fgColor='1E293B')
    GRN_FILL  = PatternFill('solid', fgColor='D1FAE5')
    YEL_FILL  = PatternFill('solid', fgColor='FEF9C3')
    RED_FILL  = PatternFill('solid', fgColor='FEE2E2')
    GRY_FILL  = PatternFill('solid', fgColor='F1F5F9')
    BLU_FILL  = PatternFill('solid', fgColor='DBEAFE')

    HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
    BOLD_FONT = Font(bold=True, size=10)
    NORM_FONT = Font(size=10)
    WARN_FONT = Font(bold=True, color='92400E', size=10)
    META_FONT = Font(size=9, color='64748B')

    THIN_SIDE = Side(style='thin', color='CBD5E1')
    BOT_BORDER = Border(bottom=THIN_SIDE)

    meses_pt = ['', 'janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
                'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
    mes_nome = f"{meses_pt[d['mes']]}/{d['ano']}"

    # ── Aba 1: Resumo Executivo ───────────────────────────────────
    ws1 = wb.create_sheet('Resumo Executivo')
    ws1.column_dimensions['A'].width = 44
    ws1.column_dimensions['B'].width = 26

    def add_titulo(text, size=13, bold=True, color='000000'):
        r = ws1.cell(row=ws1.max_row + 1 if ws1.max_row else 1, column=1, value=text)
        r.font = Font(bold=bold, size=size, color=color)
        ws1.merge_cells(
            start_row=r.row, start_column=1,
            end_row=r.row, end_column=2
        )
        return r.row

    def add_linha(label, value, bold=False, fill=None):
        row_n = (ws1.max_row or 0) + 1
        c1 = ws1.cell(row=row_n, column=1, value=label)
        c2 = ws1.cell(row=row_n, column=2, value=value)
        for c in (c1, c2):
            c.font  = BOLD_FONT if bold else NORM_FONT
            c.border = BOT_BORDER
            if fill:
                c.fill = fill

    add_titulo('APURAÇÃO PIS/COFINS — MONTANA SEGURANÇA PRIVADA LTDA')
    add_titulo(
        f'Competência: {mes_nome.upper()}  |  Regime: Lucro Real Anual — Cumulativo',
        size=9, bold=False, color='64748B'
    )
    add_titulo(
        f'Base: Regime de Caixa — créditos recebidos em {mes_nome}',
        size=9, bold=False, color='64748B'
    )
    ws1.append([])  # linha em branco

    add_linha('BASE DE CÁLCULO (créditos tributáveis no mês)', brl(d['base_tributavel']), True,  BLU_FILL)
    add_linha(f"PIS — 0,65%  (DARF {DARF_PIS})",             brl(d['pis']),              False, GRN_FILL)
    add_linha(f"COFINS — 3,00%  (DARF {DARF_COFINS})",       brl(d['cofins']),           False, GRN_FILL)
    add_linha('TOTAL A RECOLHER (PIS + COFINS)',              brl(d['total_darf']),       True,  GRN_FILL)
    add_linha('VENCIMENTO DARF',                              d['vencimento'],            True)
    ws1.append([])

    add_linha('Qtd. créditos tributáveis',              str(len(d['tributaveis'])))
    add_linha('Qtd. excluídos (NFs emitidas 2024/2025)',str(len(d['excluidos'])))
    add_linha('Qtd. não tributáveis (internos/invest.)', str(len(d['nao_tributa'])))
    add_linha(
        'Qtd. PENDENTES (classificar no Portal da Transparência)',
        str(len(d['pendentes'])),
        fill=YEL_FILL if d['pendentes'] else None
    )
    ws1.append([])

    if d['pendentes']:
        row_n = (ws1.max_row or 0) + 1
        c = ws1.cell(row=row_n, column=1,
                     value='⚠ ATENÇÃO: Há créditos sem NF vinculada. Veja a aba "Pendentes de Classificação".')
        c.font = WARN_FONT
        c.fill = YEL_FILL
        ws1.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=2)
        ws1.append([])

    add_linha('Gerado em', date.today().strftime('%d/%m/%Y'), False, GRY_FILL)

    # ── Helper: aba de detalhe ────────────────────────────────────
    def detail_sheet(name, rows, row_fill, cols):
        ws = wb.create_sheet(name)
        for idx, (key, title, width) in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
            c = ws.cell(row=1, column=idx, value=title)
            c.font      = HDR_FONT
            c.fill      = HDR_FILL
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

        MONEY_KEYS = {'credito', 'nf_valor_bruto', 'nf_valor_liquido'}
        for r in rows:
            row_data = []
            for key, _, _ in cols:
                v = r.get(key)
                row_data.append('' if v is None else v)
            row_obj = ws.append(row_data)
            for idx, (key, _, _) in enumerate(cols, 1):
                cell = ws.cell(row=ws.max_row, column=idx)
                cell.fill   = row_fill
                cell.border = BOT_BORDER
                if key in MONEY_KEYS:
                    try:
                        cell.value        = float(r.get(key) or 0)
                        cell.number_format = 'R$ #,##0.00'
                    except (ValueError, TypeError):
                        pass

    DET_COLS = [
        ('data_iso',             'Data',            12),
        ('historico',            'Histórico',        40),
        ('pagador_identificado', 'Pagador',          26),
        ('credito',              'Valor Crédito',    15),
        ('nf_numero',            'NF Nº',            10),
        ('data_emissao',         'Data Emissão NF',  15),
        ('nf_competencia',       'Competência NF',   13),
        ('tomador',              'Tomador',          22),
        ('status_conciliacao',   'Status',           12),
    ]

    detail_sheet('Créditos Tributáveis',        d['tributaveis'], GRN_FILL, DET_COLS)
    detail_sheet('Excluídos (Exerc. Anterior)', d['excluidos'],   RED_FILL, DET_COLS)
    detail_sheet('Não Tributa', d['nao_tributa'], GRY_FILL, [
        ('data_iso',  'Data',          12),
        ('historico', 'Histórico',     52),
        ('credito',   'Valor Crédito', 15),
    ])
    detail_sheet('Pendentes de Classificação', d['pendentes'], YEL_FILL, [
        ('data_iso',             'Data',          12),
        ('historico',            'Histórico',     42),
        ('pagador_identificado', 'Pagador',        26),
        ('credito',              'Valor Crédito', 15),
        ('status_conciliacao',   'Status',        12),
    ])

    wb.save(str(out_path))
    return out_path


# ─── main ─────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='Apuração PIS/COFINS — Montana Segurança (cumulativo, caixa)'
    )
    parser.add_argument('--ano-mes', required=True,
                        help='Período AAAA-MM (ex: 2026-03)')
    args = parser.parse_args()

    # Garante UTF-8 no stdout (Windows usa cp1252 por padrao)
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    ano_mes = args.ano_mes.strip()
    if len(ano_mes) != 7 or ano_mes[4] != '-':
        print("ERRO: formato invalido '" + ano_mes + "'. Use AAAA-MM (ex: 2026-03).")
        sys.exit(1)

    print("\n  Montana Seguranca - Apuracao PIS/COFINS " + ano_mes)
    print("  Banco: " + str(DB_PATH))

    if not DB_PATH.exists():
        print("\n  ERRO: Banco nao encontrado em " + str(DB_PATH))
        sys.exit(1)

    d = apurar(ano_mes)

    print("\n  --- Resultado -------------------------------------------")
    print("  Base tributavel:  " + brl(d['base_tributavel']))
    print("  PIS  (0,65%):     " + brl(d['pis']) + "    DARF " + DARF_PIS)
    print("  COFINS (3,00%):   " + brl(d['cofins']) + "    DARF " + DARF_COFINS)
    print("  TOTAL a recolher: " + brl(d['total_darf']))
    print("  Vencimento DARF:  " + d['vencimento'])
    print("\n  --- Classificacao dos creditos --------------------------")
    print("  Tributaveis:   %4d creditos" % len(d['tributaveis']))
    print("  Excluidos:     %4d creditos  (NFs emitidas 2024/2025)" % len(d['excluidos']))
    print("  Nao tributa:   %4d creditos  (internos / investimentos)" % len(d['nao_tributa']))
    print("  PENDENTES:     %4d creditos  <- classificar no Portal" % len(d['pendentes']))

    if d['pendentes']:
        print("\n  ATENCAO: %d creditos sem NF vinculada - verifique e vincule no ERP." % len(d['pendentes']))

    out = gerar_excel(d)
    print("\n  OK - Excel gerado: " + str(out) + "\n")


if __name__ == '__main__':
    main()
