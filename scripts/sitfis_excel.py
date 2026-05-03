#!/usr/bin/env python3
"""
SITFIS -> Excel consolidado (4 empresas)

Gera um .xlsx com abas:
  - Resumo        : 1 linha por empresa (certidão, totais, validade)
  - Debitos_SIEF  : todos os débitos RFB
  - Parcelamentos : SIEFPAR + SISPAR
  - Inscricoes_PGFN : SIDA (ativa + suspensa)
  - Alertas       : itens que exigem ação (CPEN vencendo, parcelas em atraso, etc.)

Uso:
  python3 sitfis_excel.py --db /opt/montana/data/sitfis.db --out /opt/montana/data/sitfis_consolidado.xlsx
"""
import argparse
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    import sys
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

EMPRESAS = {
    "14.092.519/0001-51": "MONTANA ASSESSORIA EMPRESARIAL LTDA",
    "05.143.442/0001-59": "MONTANA SEGURANÇA LTDA",
    "11.133.664/0001-10": "PORTO DO VAU SEGURANÇA LTDA",
    "17.900.439/0001-03": "MUSTANG SEGURANÇA LTDA",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALERT_FILL = PatternFill("solid", fgColor="FFF2CC")
DANGER_FILL = PatternFill("solid", fgColor="F8CBAD")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def autosize(ws):
    for col in ws.columns:
        mx = 10
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            mx = max(mx, min(60, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = mx


def money_fmt(cell):
    cell.number_format = 'R$ #,##0.00'


def br_date(iso):
    if not iso:
        return None
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return iso


def build(db_path, out_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    wb = openpyxl.Workbook()

    # ---------------- Resumo ----------------
    ws = wb.active
    ws.title = "Resumo"
    cols = [
        "CNPJ", "Razão Social", "Tipo Certidão", "Emissão", "Validade",
        "Dias até Vencer", "Código Controle", "UA Domicílio",
        "Situação Cadastral", "Débitos SIEF (R$)", "Parcelamentos Saldo (R$)",
        "Inscrições PGFN (R$)", "Qtd Débitos", "Qtd Parc.", "Qtd Inscr.",
        "Fonte",
    ]
    ws.append(cols)
    style_header(ws, len(cols))

    today = date.today()
    resumo_rows = []
    for cnpj, razao_default in EMPRESAS.items():
        snap = conn.execute(
            "SELECT * FROM sitfis_snapshots WHERE cnpj=? "
            "ORDER BY emissao DESC, id DESC LIMIT 1",
            (cnpj,),
        ).fetchone()
        if not snap:
            ws.append([cnpj, razao_default, "— sem dados —", None, None, None,
                       None, None, None, 0, 0, 0, 0, 0, 0, "aguardando SERPRO"])
            continue
        sid = snap["id"]
        total_deb = conn.execute(
            "SELECT COALESCE(SUM(saldo_consolidado),0), COUNT(*) "
            "FROM sitfis_debitos WHERE snapshot_id=?", (sid,)
        ).fetchone()
        total_parc = conn.execute(
            "SELECT COALESCE(SUM(valor_em_atraso),0), COUNT(*) "
            "FROM sitfis_parcelamentos WHERE snapshot_id=?", (sid,)
        ).fetchone()
        total_insc = conn.execute(
            "SELECT 0, COUNT(*) "
            "FROM sitfis_inscricoes WHERE snapshot_id=?", (sid,)
        ).fetchone()
        dias = None
        if snap["validade_certidao"]:
            try:
                dias = (datetime.strptime(snap["validade_certidao"], "%Y-%m-%d").date() - today).days
            except ValueError:
                pass
        row = [
            snap["cnpj"], snap["razao_social"] or razao_default,
            snap["tipo_certidao"], br_date(snap["emissao"]),
            br_date(snap["validade_certidao"]), dias,
            snap["codigo_controle"], snap["ua_domicilio"],
            snap["situacao_cadastral"],
            total_deb[0], total_parc[0], total_insc[0],
            total_deb[1], total_parc[1], total_insc[1],
            Path(snap["arquivo"]).name if snap["arquivo"] else "",
        ]
        ws.append(row)
        resumo_rows.append(row)

    # money formatting + alerts
    for r in range(2, ws.max_row + 1):
        for c in (10, 11, 12):
            money_fmt(ws.cell(row=r, column=c))
        dias_cell = ws.cell(row=r, column=6)
        if isinstance(dias_cell.value, int):
            if dias_cell.value < 0:
                for c in range(1, len(cols) + 1):
                    ws.cell(row=r, column=c).fill = DANGER_FILL
            elif dias_cell.value <= 30:
                for c in range(1, len(cols) + 1):
                    ws.cell(row=r, column=c).fill = ALERT_FILL
    autosize(ws)

    # ---------------- Debitos ----------------
    ws2 = wb.create_sheet("Debitos_SIEF")
    cols = ["CNPJ", "Razão Social", "Receita", "Período",
            "Vencimento", "Valor Original",
            "Saldo Devedor", "Multa", "Juros", "Saldo Consolidado", "Situação"]
    ws2.append(cols)
    style_header(ws2, len(cols))
    for row in conn.execute("""
        SELECT s.cnpj, s.razao_social, d.*
        FROM sitfis_debitos d
        JOIN sitfis_snapshots s ON s.id = d.snapshot_id
        ORDER BY s.cnpj, d.data_vencimento
    """):
        ws2.append([row["cnpj"], row["razao_social"], row["receita"],
                    row["periodo_apuracao"], br_date(row["data_vencimento"]),
                    row["valor_original"], row["saldo_devedor"],
                    row["multa"], row["juros"], row["saldo_consolidado"],
                    row["situacao"]])
    for r in range(2, ws2.max_row + 1):
        for c in (6, 7, 8, 9, 10):
            money_fmt(ws2.cell(row=r, column=c))
    autosize(ws2)

    # ---------------- Parcelamentos ----------------
    ws3 = wb.create_sheet("Parcelamentos")
    cols = ["CNPJ", "Razão Social", "Sistema", "Número/Conta",
            "Modalidade", "Parcelas em Atraso", "Valor em Atraso",
            "Situação", "Exigibilidade Suspensa"]
    ws3.append(cols)
    style_header(ws3, len(cols))
    for row in conn.execute("""
        SELECT s.cnpj, s.razao_social, p.*
        FROM sitfis_parcelamentos p
        JOIN sitfis_snapshots s ON s.id = p.snapshot_id
        ORDER BY s.cnpj, p.sistema
    """):
        ws3.append([row["cnpj"], row["razao_social"], row["sistema"],
                    row["numero"] or row["conta"],
                    row["modalidade"], row["parcelas_em_atraso"],
                    row["valor_em_atraso"], row["situacao"],
                    "Sim" if row["exigibilidade_suspensa"] else "Não"])
    for r in range(2, ws3.max_row + 1):
        money_fmt(ws3.cell(row=r, column=7))
        if ws3.cell(row=r, column=6).value and isinstance(ws3.cell(row=r, column=6).value, int) and ws3.cell(row=r, column=6).value > 0:
            for c in range(1, len(cols) + 1):
                ws3.cell(row=r, column=c).fill = DANGER_FILL
    autosize(ws3)

    # ---------------- Inscricoes ----------------
    ws4 = wb.create_sheet("Inscricoes_PGFN")
    cols = ["CNPJ", "Razão Social", "Inscrição", "Receita",
            "Inscrito em", "Ajuizado em", "Processo",
            "Tipo Devedor", "Situação", "Exigibilidade Suspensa"]
    ws4.append(cols)
    style_header(ws4, len(cols))
    for row in conn.execute("""
        SELECT s.cnpj, s.razao_social, i.*
        FROM sitfis_inscricoes i
        JOIN sitfis_snapshots s ON s.id = i.snapshot_id
        ORDER BY s.cnpj, i.inscrito_em
    """):
        ws4.append([row["cnpj"], row["razao_social"], row["inscricao"],
                    row["receita"], br_date(row["inscrito_em"]),
                    br_date(row["ajuizado_em"]), row["processo"],
                    row["tipo_devedor"], row["situacao"],
                    "Sim" if row["exigibilidade_suspensa"] else "Não"])
    autosize(ws4)

    # ---------------- Alertas ----------------
    ws5 = wb.create_sheet("Alertas")
    cols = ["Severidade", "CNPJ", "Razão Social", "Tipo", "Descrição", "Valor (R$)"]
    ws5.append(cols)
    style_header(ws5, len(cols))

    # CPEN vencendo
    for snap in conn.execute("SELECT * FROM sitfis_snapshots"):
        if not snap["validade_certidao"]:
            continue
        try:
            v = datetime.strptime(snap["validade_certidao"], "%Y-%m-%d").date()
        except ValueError:
            continue
        dias = (v - today).days
        if dias < 0:
            ws5.append(["CRÍTICO", snap["cnpj"], snap["razao_social"],
                        "Certidão Vencida",
                        f"{snap['tipo_certidao']} venceu em {br_date(snap['validade_certidao'])} ({-dias}d atrás)",
                        None])
        elif dias <= 30:
            ws5.append(["ATENÇÃO", snap["cnpj"], snap["razao_social"],
                        "Certidão Vencendo",
                        f"{snap['tipo_certidao']} vence em {dias} dias ({br_date(snap['validade_certidao'])})",
                        None])

    # Parcelamentos em atraso
    for row in conn.execute("""
        SELECT s.cnpj, s.razao_social, p.sistema, p.numero, p.conta,
               p.parcelas_em_atraso, p.valor_em_atraso
        FROM sitfis_parcelamentos p
        JOIN sitfis_snapshots s ON s.id = p.snapshot_id
        WHERE p.parcelas_em_atraso > 0
    """):
        ws5.append(["CRÍTICO", row["cnpj"], row["razao_social"],
                    "Parcelamento em Atraso",
                    f"{row['sistema']} {row['numero'] or row['conta']}: "
                    f"{row['parcelas_em_atraso']} parc.",
                    row["valor_em_atraso"]])

    # Novas inscrições PGFN (últimos 90 dias)
    cutoff = (today - timedelta(days=90)).isoformat()
    for row in conn.execute("""
        SELECT s.cnpj, s.razao_social, i.inscricao, i.inscrito_em, i.situacao
        FROM sitfis_inscricoes i
        JOIN sitfis_snapshots s ON s.id = i.snapshot_id
        WHERE i.inscrito_em >= ? AND i.exigibilidade_suspensa = 0
    """, (cutoff,)):
        ws5.append(["ATENÇÃO", row["cnpj"], row["razao_social"],
                    "Nova Inscrição PGFN",
                    f"{row['inscricao']} inscrita em {br_date(row['inscrito_em'])} "
                    f"({row['situacao']})",
                    None])

    # Débitos SIEF abertos
    for row in conn.execute("""
        SELECT s.cnpj, s.razao_social, d.receita, d.periodo_apuracao,
               d.saldo_consolidado
        FROM sitfis_debitos d
        JOIN sitfis_snapshots s ON s.id = d.snapshot_id
        WHERE d.situacao = 'DEVEDOR'
    """):
        ws5.append(["ATENÇÃO", row["cnpj"], row["razao_social"],
                    "Débito RFB em Aberto",
                    f"{row['receita']} {row['periodo_apuracao']}",
                    row["saldo_consolidado"]])

    for r in range(2, ws5.max_row + 1):
        money_fmt(ws5.cell(row=r, column=6))
        sev = ws5.cell(row=r, column=1).value
        if sev == "CRÍTICO":
            for c in range(1, len(cols) + 1):
                ws5.cell(row=r, column=c).fill = DANGER_FILL
        elif sev == "ATENÇÃO":
            for c in range(1, len(cols) + 1):
                ws5.cell(row=r, column=c).fill = ALERT_FILL
    autosize(ws5)

    os_path = Path(out_path)
    os_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(os_path)
    conn.close()
    print(f"[ok] gravado: {out_path}")
    print(f"     abas: Resumo, Debitos_SIEF, Parcelamentos, Inscricoes_PGFN, Alertas")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default="/opt/montana/data/sitfis.db")
    ap.add_argument("--out", default="/opt/montana/data/sitfis_consolidado.xlsx")
    args = ap.parse_args()
    build(args.db, args.out)


if __name__ == "__main__":
    main()
