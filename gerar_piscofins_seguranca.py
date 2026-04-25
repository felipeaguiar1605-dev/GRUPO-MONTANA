#!/usr/bin/env python3
"""Gera relatório PIS/COFINS Montana Segurança Mar/2026"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT = '/sessions/confident-lucid-rubin/mnt/app_unificado/Apuracao_PISCOFINS_Seguranca_202603.xlsx'

AZUL_ESC  = '1F3864'; AZUL_MED = '2E75B6'; AZUL_CLA = 'D6E4F0'
VERDE_ESC = '1B5E20'; VERDE_CLA = 'E8F5E9'; AMBAR_CLA = 'FFF8E1'
CINZA_CLA = 'F5F5F5'; LARANJA_CLA = 'FFF3CD'; VERMELHO_CLA = 'FFEBEE'
BRANCO = 'FFFFFF'

def fill(h): return PatternFill('solid', start_color=h, fgColor=h)
def font(bold=False, color='000000', size=11, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
def bt():
    s = Side(style='thin', color='BDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def right():  return Alignment(horizontal='right',  vertical='center')
def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)

FMT_BRL = '#,##0.00'; FMT_PCT = '0.00%'

# ─────────────────────────────────────────────────────────────────
# DADOS
# ─────────────────────────────────────────────────────────────────

# BB - 49 créditos únicos CONCILIADO (após dedup) = serviços recebidos
BB_CONCILIADO = [
    ('18/03/2026', 193359.33, 'GOVERNO DO ESTADO TO',          '01786029000103', 'TED GOVERNO DO EST'),
    ('19/03/2026', 147341.41, 'MUNICIPIO DE PALMAS',           '24851511004849', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('19/03/2026', 122784.51, 'MUNICIPIO DE PALMAS',           '24851511002200', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('19/03/2026',  99390.38, 'MUNICIPIO DE PALMAS',           '24851511001408', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('19/03/2026',  73670.71, 'MUNICIPIO DE PALMAS',           '24851511001904', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('12/03/2026',  56545.49, 'GOVERNO DO ESTADO TO',          '01786029000103', 'TED GOVERNO DO EST'),
    ('03/03/2026',  51013.11, 'FUNDACAO UFT',                  '05149726000104', 'Pix FUNDACAO UN'),
    ('17/03/2026',  49113.81, 'MUNICIPIO DE PALMAS',           '24851511003605', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('26/03/2026',  49113.81, 'MUNICIPIO DE PALMAS',           '24851511000932', 'Ordem Bancária ORDENS BANCARIAS'),
    ('17/03/2026',  49113.80, 'MUNICIPIO DE PALMAS',           '24851511003524', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('03/03/2026',  42618.71, 'FUNDACAO UFT',                  '05149726000104', 'Pix FUNDACAO UN'),
    ('03/03/2026',  41933.74, 'FUNDACAO UFT',                  '05149726000104', 'Pix FUNDACAO UN'),
    ('03/03/2026',  39653.99, 'FUNDACAO UFT',                  '05149726000104', 'Pix FUNDACAO UN'),
    ('17/03/2026',  36016.79, 'MUNICIPIO DE PALMAS',           '24851511000428', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',  35862.16, 'MUNICIPIO DE PALMAS',           '24851511002986', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('19/03/2026',  35862.16, 'MUNICIPIO DE PALMAS',           '24851511001076', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('06/03/2026',  28474.95, 'GOVERNO DO ESTADO TO',          '01786029000103', 'TED GOVERNO DO EST'),
    ('16/03/2026',  28474.95, 'GOVERNO DO ESTADO TO',          '01786029000103', 'TED GOVERNO DO EST'),
    ('19/03/2026',  26194.03, 'MUNICIPIO DE PALMAS',           '24851511001238', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',  24556.90, 'MUNICIPIO DE PALMAS',           '24851511004415', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('19/03/2026',  24556.90, 'MUNICIPIO DE PALMAS',           '24851511000770', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('26/03/2026',  24556.90, 'MUNICIPIO DE PALMAS',           '24851511004849', 'Ordem Bancária ORDENS BANCARIAS'),
    ('03/03/2026',  20749.12, 'FUNDACAO UFT',                  '05149726000104', 'Pix FUNDACAO UN'),
    ('16/03/2026',  19645.52, 'MUNICIPIO DE PALMAS',           '24851511001904', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('18/03/2026',  19038.27, 'MUNICIPIO DE PALMAS',           '24851511004849', 'Ordem Bancária ORDENS BANCARIAS'),
    ('06/03/2026',  18228.77, 'GOVERNO DO ESTADO TO',          '01786029000103', 'TED GOVERNO DO EST'),
    ('16/03/2026',  18008.39, 'MUNICIPIO DE PALMAS',           '24851511003605', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',  13097.01, 'MUNICIPIO DE PALMAS',           '24851511001904', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',  12859.84, 'MUNICIPIO DE PALMAS',           '24851511001904', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('26/03/2026',  12859.84, 'MUNICIPIO DE PALMAS',           '24851511001904', 'Ordem Bancária ORDENS BANCARIAS'),
    ('26/03/2026',  11697.07, 'MUNICIPIO DE PALMAS',           '24851511002200', 'Ordem Bancária ORDENS BANCARIAS'),
    ('26/03/2026',   9667.72, 'MUNICIPIO DE PALMAS',           '24851511001408', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   9563.24, 'MUNICIPIO DE PALMAS',           '24851511003524', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('06/03/2026',   8312.40, 'GOVERNO DO ESTADO TO',          '01786029000103', 'TED GOVERNO DO EST'),
    ('16/03/2026',   8185.64, 'MUNICIPIO DE PALMAS',           '24851511002200', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   8185.63, 'MUNICIPIO DE PALMAS',           '24851511000932', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   6548.50, 'MUNICIPIO DE PALMAS',           '24851511002986', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   6139.23, 'MUNICIPIO DE PALMAS',           '24851511003605', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   6139.22, 'MUNICIPIO DE PALMAS',           '24851511001408', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   6002.80, 'MUNICIPIO DE PALMAS',           '24851511001904', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   4502.11, 'MUNICIPIO DE PALMAS',           '24851511003605', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   4502.10, 'MUNICIPIO DE PALMAS',           '24851511001408', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   3429.28, 'MUNICIPIO DE PALMAS',           '24851511004849', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   3274.25, 'MUNICIPIO DE PALMAS',           '24851511003524', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('11/03/2026',   3262.39, 'Transferência recebida',        '',               'Transferência ISAIAS OLIVEIRA PEREIRA'),
    ('16/03/2026',   2182.84, 'MUNICIPIO DE PALMAS',           '24851511000770', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   2182.83, 'MUNICIPIO DE PALMAS',           '24851511000428', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('16/03/2026',   1637.13, 'MUNICIPIO DE PALMAS',           '24851511004849', 'Ordem Bancária MUNICIPIO DE PALMAS'),
    ('17/03/2026',   1637.13, 'MUNICIPIO DE PALMAS',           '24851511001904', 'Ordem Bancária MUNICIPIO DE PALMAS'),
]

BB_INVESTIMENTO = [
    ('06/03/2026', 796425.09, 'BB Rende Fácil / Rende Facil'),
    ('20/03/2026', 554266.60, 'BB Rende Fácil / Rende Facil'),
    ('24/03/2026', 214320.55, 'BB Rende Fácil'),
    ('19/03/2026', 142985.15, 'BB Rende Fácil / Rende Facil'),
    ('04/03/2026',  41276.92, 'BB Rende Fácil / Rende Facil'),
    ('02/03/2026',  19925.23, 'BB Rende Fácil / Rende Facil'),
    ('23/03/2026',  12919.15, 'BB Rende Fácil / Rende Facil'),
    ('05/03/2026',  12204.19, 'BB Rende Fácil / Rende Facil'),
    ('10/03/2026',   3999.29, 'BB Rende Fácil / Rende Facil'),
    ('13/03/2026',   3895.90, 'BB Rende Fácil / Rende Facil'),
]

BB_INTERNO = [
    ('06/03/2026', 320000.00, 'PIX MONTANA SEGURANCA PR (CNPJ 19200109000109)'),
    ('09/03/2026',  50000.00, 'Transferência MONTANA SERVICOS'),
    ('11/03/2026',  30000.00, 'Transferência MONTANA SERVICOS'),
    ('06/03/2026',  10000.00, 'Transferência MONTANA SERVICOS'),
    ('06/03/2026',   5000.00, 'Transferência MONTANA SERVICOS'),
]

BRB_INVESTIMENTO = [
    ('02/03/2026',  47500.00, 'RESGATE CDB/RDB — BRB 031.015.474-0'),
    ('03/03/2026',  50000.00, 'RESGATE CDB/RDB — BRB 031.015.474-0'),
    ('03/03/2026',  47500.00, 'RESGATE CDB/RDB — BRB 031.015.474-0'),
    ('05/03/2026', 120000.00, 'RESG FI BRB FEDERAL INVEST DOC 160512'),
    ('06/03/2026',      6.12, 'CRED JUROS CDB AUTOMAT — BRB'),
    ('06/03/2026', 320000.00, 'RESGATE CDB/RDB — BRB 031.015.474-0'),
    ('09/03/2026',     19.21, 'CRED JUROS CDB AUTOMAT — BRB'),
    ('09/03/2026',  20000.00, 'RESG FI BRB FEDERAL INVEST DOC 160508'),
    ('23/03/2026', 523785.55, 'RESGATE CDB/RDB — BRB 031.015.474-0'),
    ('24/03/2026',  50000.01, 'RESGATE CDB/RDB — BRB 031.015.474-0'),
    ('26/03/2026', 300000.00, 'RESGATE CDB/RDB — BRB 031.015.474-0'),
]

# NFs PENDENTE por tomador (DIFERIDO — regime caixa)
NFS_PENDENTE = [
    ('MINISTERIO PUBLICO DO ESTADO DO TOCANTINS', '01786078000146', 111, 699738.40),
    ('UNIVERSIDADE ESTADUAL DO TOCANTINS - UNITINS','01637536000185',10, 529134.32),
    ('MUNICIPIO DE PALMAS',                        '24851511002200',  3, 508501.52),
    ('FUNDACAO UNIVERSIDADE FEDERAL DO TOCANTINS', '05149726000104', 10, 728099.08),
    ('MUNICIPIO DE PALMAS',                        '24851511000428',  2, 376208.44),
    ('FUNDACAO CULTURAL DE PALMAS - FCP',          '11794886000109',  3, 343013.95),
    ('MUNICIPIO DE PALMAS',                        '24851511001904',  4, 266158.02),
    ('DEPARTAMENTO ESTADUAL DE TRANSITO',          '26752857000151',  5, 235690.48),
    ('MUNICIPIO DE PALMAS',                        '24851511003524',  2, 124024.76),
    ('SECRETARIA DA EDUCACAO',                     '25053083000108',  2, 106287.20),
    ('TRIBUNAL DE CONTAS DO ESTADO DO TOCANTINS',  '25053133000157',  1,  71395.82),
    ('CORPO DE BOMBEIROS MILITAR DO ESTADO DO TO', '07924551000190',  1,  35953.21),
    ('MUNICIPIO DE PALMAS',                        '24851511004849',  1,  33205.69),
    ('INSTITUTO DE PREVIDENCIA SOCIAL DE PALMAS',  '05278848000109',  1,  24038.23),
    ('MUNICIPIO DE PALMAS',                        '24851511000770',  1,  16536.63),
]

# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────
def scw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def hrow(ws, row, cols, texts, bg, fg='FFFFFF', sz=11, bold=True):
    for c, t in zip(cols, texts):
        cell = ws.cell(row, c, t)
        cell.font = font(bold, fg, sz)
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = bt()

def dc(ws, row, col, val, fmt=None, bold=False, bg=None, al=None):
    cell = ws.cell(row, col, val)
    cell.font = font(bold)
    cell.border = bt()
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = fill(bg)
    cell.alignment = al or left()
    return cell

def money(ws, row, col, val, bold=False, bg=None):
    dc(ws, row, col, val, FMT_BRL, bold, bg, right())

def title_block(ws, row, text, sub=None, ncols=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = Font(name='Arial', bold=True, color=BRANCO, size=14)
    c.fill = fill(AZUL_ESC); c.alignment = center()
    if sub:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=ncols)
        c2 = ws.cell(row+1, 1, sub)
        c2.font = Font(name='Arial', italic=True, color=AZUL_MED, size=10)
        c2.fill = fill(AZUL_CLA); c2.alignment = center()

# ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════ ABA 1: RESUMO EXECUTIVO ══════════════════════════
ws1 = wb.active; ws1.title = 'Resumo Executivo'
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40; ws1.row_dimensions[2].height = 22
for c, w in enumerate([3,30,22,20,18,18,18,10], 1): scw(ws1, c, w)

title_block(ws1, 1,
    'APURAÇÃO PIS/COFINS — MARÇO/2026',
    'Montana Segurança Privada Ltda  ·  Lucro Real Não-Cumulativo  ·  Regime de Caixa  ·  ' + datetime.now().strftime('%d/%m/%Y'))

base_trib   = sum(v for _,v,*_ in BB_CONCILIADO)
bb_inv      = sum(v for _,v,_ in BB_INVESTIMENTO)
bb_int      = sum(v for _,v,_ in BB_INTERNO)
brb_inv     = sum(v for _,v,_ in BRB_INVESTIMENTO)
nf_dif      = sum(v for *_,v in NFS_PENDENTE)
total_ext   = base_trib + bb_inv + bb_int + brb_inv + 883.52  # +PENDENTE extrato

pis_deb     = base_trib * 0.0165
cofins_deb  = base_trib * 0.076
total_bruto = pis_deb + cofins_deb

# bloco composição
r = 4
ws1.row_dimensions[r].height = 20
hrow(ws1, r, [2,3,4,5], ['COMPOSIÇÃO DOS CRÉDITOS','Qtd','Valor (R$)','% Total'], AZUL_MED, sz=10)
r += 1

comp = [
    ('BB — Créditos Tributáveis (CONCILIADO)',  49,  base_trib, True,  VERDE_CLA),
    ('BB — Rende Fácil (INVESTIMENTO)',         10,  bb_inv,    False, CINZA_CLA),
    ('BB — Transferências Internas (INTERNO)',   5,  bb_int,    False, CINZA_CLA),
    ('BRB — Resgates CDB/FI BRB (INVESTIMENTO)',11,  brb_inv,   False, CINZA_CLA),
    ('BB — PENDENTE (pequenos)',                 2,  883.52,    False, CINZA_CLA),
    ('TOTAL GERAL',                             77,  base_trib+bb_inv+bb_int+brb_inv+883.52, True, AZUL_CLA),
]
total_row_comp = r + len(comp) - 1
for label, qtd, val, bold, bg in comp:
    ws1.cell(r,2,label).font = font(bold=bold); ws1.cell(r,2).fill=fill(bg); ws1.cell(r,2).border=bt(); ws1.cell(r,2).alignment=left()
    ws1.cell(r,3,qtd).number_format='#,##0'; ws1.cell(r,3).fill=fill(bg); ws1.cell(r,3).border=bt(); ws1.cell(r,3).alignment=center(); ws1.cell(r,3).font=font(bold=bold)
    ws1.cell(r,4,val).number_format=FMT_BRL; ws1.cell(r,4).fill=fill(bg); ws1.cell(r,4).border=bt(); ws1.cell(r,4).alignment=right(); ws1.cell(r,4).font=font(bold=bold)
    ws1.cell(r,5).fill=fill(bg); ws1.cell(r,5).border=bt()
    r += 1

# bloco apuração
r += 1
ws1.row_dimensions[r].height = 20
hrow(ws1, r, [2,3,4,5], ['APURAÇÃO DO IMPOSTO','','Valor (R$)',''], AZUL_ESC, sz=11)
r += 1

apuracao = [
    ('BASE TRIBUTÁVEL (BB CONCILIADO)',              base_trib,    True,  VERDE_CLA,  False),
    ('',                                             None,         False, BRANCO,     False),
    ('PIS — débito (1,65%)',                         pis_deb,      False, CINZA_CLA,  False),
    ('  ⚠  (-) Crédito PIS — a apurar com contador', None,        False, AMBAR_CLA,  False),
    ('PIS A RECOLHER (bruto s/ créditos)',            pis_deb,     True,  AMBAR_CLA,  True),
    ('',                                             None,         False, BRANCO,     False),
    ('COFINS — débito (7,60%)',                       cofins_deb,  False, CINZA_CLA,  False),
    ('  ⚠  (-) Crédito COFINS — a apurar com contador',None,      False, AMBAR_CLA,  False),
    ('COFINS A RECOLHER (bruto s/ créditos)',          cofins_deb, True,  AMBAR_CLA,  True),
    ('',                                             None,         False, BRANCO,     False),
    ('TOTAL PIS + COFINS (BRUTO)',                    total_bruto, True,  AZUL_ESC,   True),
]
for label, val, bold, bg, white_text in apuracao:
    ws1.cell(r,2,label).font      = font(bold=bold, color=BRANCO if white_text else '000000')
    ws1.cell(r,2).fill            = fill(bg); ws1.cell(r,2).border=bt(); ws1.cell(r,2).alignment=left()
    ws1.cell(r,3).fill=fill(bg); ws1.cell(r,3).border=bt()
    ws1.cell(r,5).fill=fill(bg); ws1.cell(r,5).border=bt()
    if val is not None:
        c = ws1.cell(r,4,val)
        c.number_format=FMT_BRL; c.fill=fill(bg); c.border=bt(); c.alignment=right()
        c.font=font(bold=bold, color=BRANCO if white_text else '000000')
    else:
        ws1.cell(r,4).fill=fill(bg); ws1.cell(r,4).border=bt()
    r += 1

# DARF
r += 1
ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws1.cell(r,2,'⚠  DARF — Vencimento: 27/04/2026  ·  Deduzir créditos de entrada antes de emitir')
c.font=Font(name='Arial',bold=True,color='7B3F00',size=10)
c.fill=fill('FFF3CD'); c.alignment=center()
for col in [2,3,4,5]: ws1.cell(r,col).border=bt()
r += 1
for lbl, val in [('PIS — cód. DARF 6912 (bruto)', pis_deb), ('COFINS — cód. DARF 5856 (bruto)', cofins_deb)]:
    ws1.cell(r,2,lbl).font=font(bold=True); ws1.cell(r,2).fill=fill(AMBAR_CLA); ws1.cell(r,2).border=bt(); ws1.cell(r,2).alignment=left()
    ws1.cell(r,4,val).number_format=FMT_BRL; ws1.cell(r,4).font=font(bold=True); ws1.cell(r,4).fill=fill(AMBAR_CLA); ws1.cell(r,4).border=bt(); ws1.cell(r,4).alignment=right()
    for col in [3,5]: ws1.cell(r,col).fill=fill(AMBAR_CLA); ws1.cell(r,col).border=bt()
    r += 1

# DIFERIDO (NFs pendentes)
r += 1
ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws1.cell(r,2,f'NFs PENDENTE (DIFERIDO — regime caixa): 157 NFs | R$ {nf_dif:,.2f} — tributar no mês do recebimento')
c.font=Font(name='Arial',bold=True,color=AZUL_MED,size=10)
c.fill=fill(AZUL_CLA); c.alignment=center()
for col in [2,3,4,5]: ws1.cell(r,col).border=bt()

# ══════════════ ABA 2: CRÉDITOS TRIBUTÁVEIS ══════════════════════
ws2 = wb.create_sheet('Créditos Tributáveis')
ws2.sheet_view.showGridLines = False
for c, w in enumerate([3,6,14,18,38,22,40,10], 1): scw(ws2, c, w)
ws2.row_dimensions[1].height = 36; ws2.row_dimensions[2].height = 18
title_block(ws2, 1,
    'CRÉDITOS TRIBUTÁVEIS BB — MARÇO/2026',
    f'49 lançamentos únicos  ·  Base: R$ {base_trib:,.2f}  ·  PIS 1,65% + COFINS 7,60% = 9,25%')

hrow(ws2, 3, [2,3,4,5,6,7,8],
    ['#','Data','Valor (R$)','CNPJ Pagador','Pagador / Tomador','Contrato / Ref.','Histórico BB'],
    AZUL_MED, sz=10)

r = 4
for i, (data, val, tom, cnpj, hist) in enumerate(BB_CONCILIADO, 1):
    bg = CINZA_CLA if i%2==0 else BRANCO
    ws2.cell(r,2,i).fill=fill(bg); ws2.cell(r,2).border=bt(); ws2.cell(r,2).alignment=center(); ws2.cell(r,2).font=font()
    ws2.cell(r,3,data).fill=fill(bg); ws2.cell(r,3).border=bt(); ws2.cell(r,3).alignment=center(); ws2.cell(r,3).font=font()
    money(ws2,r,4,val,bg=bg)
    ws2.cell(r,5,cnpj).fill=fill(bg); ws2.cell(r,5).border=bt(); ws2.cell(r,5).alignment=center(); ws2.cell(r,5).font=font(size=9)
    ws2.cell(r,6,tom).fill=fill(bg); ws2.cell(r,6).border=bt(); ws2.cell(r,6).alignment=left(); ws2.cell(r,6).font=font()
    ws2.cell(r,7,'Contrato BB - serviços prestados').fill=fill(bg); ws2.cell(r,7).border=bt(); ws2.cell(r,7).alignment=left(); ws2.cell(r,7).font=font(size=9)
    ws2.cell(r,8,hist[:50]).fill=fill(bg); ws2.cell(r,8).border=bt(); ws2.cell(r,8).alignment=left(); ws2.cell(r,8).font=font(size=9)
    r += 1
hrow(ws2, r, [2,3,4,5,6,7,8],
    ['TOTAL','49',base_trib,'','','',''], AZUL_ESC)
ws2.cell(r,4).number_format=FMT_BRL; ws2.cell(r,3).alignment=center()

# ══════════════ ABA 3: NÃO TRIBUTA ═══════════════════════════════
ws3 = wb.create_sheet('Não Tributa')
ws3.sheet_view.showGridLines = False
for c, w in enumerate([3,12,45,20,20], 1): scw(ws3, c, w)
ws3.row_dimensions[1].height = 36; ws3.row_dimensions[2].height = 18
title_block(ws3, 1,
    'CRÉDITOS NÃO TRIBUTÁVEIS — MARÇO/2026',
    'BB (Rende Fácil + Interno) + BRB (CDB/FI BRB)  ·  NÃO integram a base de PIS/COFINS')

hrow(ws3, 3, [2,3,4,5], ['Data','Banco / Categoria','Valor (R$)','Classificação'], AZUL_MED, sz=10)

nt_items = []
for d,v,h in BB_INVESTIMENTO:   nt_items.append((d,'BB — Rende Fácil',v,'INVESTIMENTO — Resgate aplicação financeira'))
for d,v,h in BB_INTERNO:        nt_items.append((d,'BB — Transferência Interna Montana',v,'INTERNO — Movimentação intra-grupo'))
for d,v,h in BRB_INVESTIMENTO:  nt_items.append((d,'BRB — CDB / FI BRB',v,'INVESTIMENTO — Resgate CDB/RDB/Fundo'))

nt_items.sort(key=lambda x: x[0][-4:]+x[0][3:5]+x[0][:2])
r = 4
total_nt = 0
for i, (data, cat, val, motivo) in enumerate(nt_items, 1):
    bg = CINZA_CLA if i%2==0 else BRANCO
    ws3.cell(r,2,data).fill=fill(bg); ws3.cell(r,2).border=bt(); ws3.cell(r,2).alignment=center(); ws3.cell(r,2).font=font()
    ws3.cell(r,3,cat).fill=fill(bg); ws3.cell(r,3).border=bt(); ws3.cell(r,3).alignment=left(); ws3.cell(r,3).font=font()
    money(ws3,r,4,val,bg=bg)
    ws3.cell(r,5,motivo).fill=fill(bg); ws3.cell(r,5).border=bt(); ws3.cell(r,5).alignment=left(); ws3.cell(r,5).font=font(size=9,italic=True)
    total_nt += val
    r += 1
hrow(ws3, r, [2,3,4,5], ['TOTAL NÃO TRIBUTÁVEL','',f'=SUM(D4:D{r-1})',''], AZUL_ESC)
ws3.cell(r,4).number_format=FMT_BRL

# ══════════════ ABA 4: DIFERIDO (NFs pendentes) ══════════════════
ws4 = wb.create_sheet('Diferido - NFs Pendentes')
ws4.sheet_view.showGridLines = False
for c, w in enumerate([3,45,22,12,22,18], 1): scw(ws4, c, w)
ws4.row_dimensions[1].height = 36; ws4.row_dimensions[2].height = 18
title_block(ws4, 1,
    f'NFs PENDENTES — DIFERIDO — MARÇO/2026',
    f'157 NFs  ·  R$ {nf_dif:,.2f}  ·  Tributar no mês do efetivo recebimento (regime caixa)')

hrow(ws4, 3, [2,3,4,5,6], ['Tomador','CNPJ','Qtd NFs','Total (R$)','Observação'], AZUL_MED, sz=10)

r = 4
for i, (tom, cnpj, qtd, val) in enumerate(NFS_PENDENTE, 1):
    bg = VERMELHO_CLA if i%2==0 else BRANCO
    ws4.cell(r,2,tom).fill=fill(bg); ws4.cell(r,2).border=bt(); ws4.cell(r,2).alignment=left(); ws4.cell(r,2).font=font(bold=True)
    ws4.cell(r,3,cnpj).fill=fill(bg); ws4.cell(r,3).border=bt(); ws4.cell(r,3).alignment=center(); ws4.cell(r,3).font=font(size=9)
    ws4.cell(r,4,qtd).fill=fill(bg); ws4.cell(r,4).border=bt(); ws4.cell(r,4).alignment=center(); ws4.cell(r,4).number_format='#,##0'; ws4.cell(r,4).font=font()
    money(ws4,r,5,val,bg=bg)
    ws4.cell(r,6,'Aguarda pagamento — tributar na competência do recebimento').fill=fill(bg); ws4.cell(r,6).border=bt(); ws4.cell(r,6).alignment=left(); ws4.cell(r,6).font=font(size=9,italic=True)
    r += 1
hrow(ws4, r, [2,3,4,5,6], ['TOTAL DIFERIDO','',157,f'=SUM(E4:E{r-1})',''], AZUL_ESC)
ws4.cell(r,5).number_format=FMT_BRL

wb.save(OUTPUT)
print(f'Salvo: {OUTPUT}')
print(f'Base tributável: R$ {base_trib:,.2f}')
print(f'PIS bruto: R$ {pis_deb:,.2f} | COFINS bruto: R$ {cofins_deb:,.2f} | Total: R$ {total_bruto:,.2f}')
