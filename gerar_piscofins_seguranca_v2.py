#!/usr/bin/env python3
"""
Montana Segurança — Apuração PIS/COFINS Março/2026
Relatório com Regra de Transição Competência → Caixa (jan/2026)

Base tributável: créditos BB CONCILIADO correspondentes a NFs de 2026
Excluídos: créditos referentes a NFs de 2024/2025 (já tributadas sob competência)
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             numbers as num_fmt)
from openpyxl.utils import get_column_letter
from datetime import datetime

OUT = 'Apuracao_PISCOFINS_Seguranca_202603_FINAL.xlsx'

# ─── PALETA ────────────────────────────────────────────────────────────────
C_VERDE_ESC  = '1B5E20'
C_VERDE_MED  = '2E7D32'
C_VERDE_CLAR = 'C8E6C9'
C_VERDE_ROWS = 'E8F5E9'
C_AMARELO    = 'FFF9C4'
C_LARANJA    = 'FFE0B2'
C_VERMELHO   = 'FFCDD2'
C_CINZA_CLAR = 'F5F5F5'
C_CINZA_MED  = 'EEEEEE'
C_AZUL_CLAR  = 'E3F2FD'
C_EXCLUIDO   = 'FCE4EC'   # rosa claro — exercício anterior excluído

def fill(hex_): return PatternFill('solid', fgColor=hex_)
def font(bold=False, color='000000', sz=10):
    return Font(bold=bold, color=color, size=sz)
def border_thin():
    s = Side(style='thin', color='BDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)
def border_med():
    s = Side(style='medium', color='757575')
    return Border(left=s, right=s, top=s, bottom=s)
def money_fmt(): return '#,##0.00'
def pct_fmt():   return '0.00%'

def set_col(ws, col, width): ws.column_dimensions[get_column_letter(col)].width = width
def merge(ws, r1, c1, r2, c2): ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def header_cell(ws, row, col, text, bg=None, bold=True, color='FFFFFF', sz=10, align='center'):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=color, size=sz)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if bg: c.fill = fill(bg)
    c.border = border_thin()
    return c

def data_cell(ws, row, col, value, fmt=None, bg=None, bold=False, align='left', color='000000'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center')
    if fmt: c.number_format = fmt
    if bg: c.fill = fill(bg)
    c.border = border_thin()
    return c

# ─── DADOS ─────────────────────────────────────────────────────────────────

# 48 créditos TRIBUTÁVEIS (todos correspondentes a NFs de 2026)
# Formato: (data_br, data_iso, valor, tomador, nf_ref, comp_nf, historico_banco)
TRIBUTAVEIS = [
    # --- Confirmados via extrato_id no ERP ---
    ('03/03/2026','2026-03-03', 20749.12,'Ministério Público do TO (MPTO)',    '202600000000107','fev/26','Pix — CNPJ 05.149.726/0001-04'),
    ('03/03/2026','2026-03-03', 39653.99,'Município de Palmas',                '202600000000304','mar/26','Pix — CNPJ 05.149.726/0001-04'),
    ('03/03/2026','2026-03-03', 42618.71,'Município de Palmas',                '202600000000289','mar/26','Pix — CNPJ 05.149.726/0001-04'),
    ('11/03/2026','2026-03-11',  3262.39,'Ministério Público do TO (MPTO)',    '202600000000143','fev/26','Transferência recebida'),
    ('16/03/2026','2026-03-16',  3274.25,'Ministério Público do TO (MPTO)',    '202600000000167','fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  6002.80,'Ministério Público do TO (MPTO)',    '202600000000173','fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  6548.50,'Ministério Público do TO (MPTO)',    '202600000000045','jan/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16', 13097.01,'Ministério Público do TO (MPTO)',    '202600000000160','fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('19/03/2026','2026-03-19', 99390.38,'Fundação Cultural de Palmas (FCP)',  '191',            'fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('19/03/2026','2026-03-19',122784.51,'Fundação Cultural de Palmas (FCP)',  '202600000000191','fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('19/03/2026','2026-03-19',147341.41,'Município de Palmas',                '202600000000233','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('26/03/2026','2026-03-26', 11697.07,'Município de Palmas',                '202600000000294','mar/26','Ordem Bancária — ORDENS BANCARIAS'),
    # --- Sem extrato_id mas melhor match = NF 2026 ---
    ('03/03/2026','2026-03-03', 41933.74,'Município de Palmas',                '187',            'fev/26','Pix — CNPJ 05.149.726/0001-04 ⚠ verificar NF'),
    ('03/03/2026','2026-03-03', 51013.11,'DETRAN-TO',                          '202600000000315','mar/26','Pix — CNPJ 05.149.726/0001-04 ⚠ verificar NF'),
    ('06/03/2026','2026-03-06',  8312.40,'Município de Palmas (ARCAF)',         '202600000000307','mar/26','TED — GOVERNO DO EST 01.786.029/0001-03'),
    ('06/03/2026','2026-03-06', 18228.77,'SEDUC-TO',                           '202600000000088','jan/26','TED — GOVERNO DO EST 01.786.029/0001-03'),
    ('06/03/2026','2026-03-06', 28474.95,'CBMTO — Corpo de Bombeiros',         '202600000000217','mar/26','TED — GOVERNO DO EST 01.786.029/0001-03'),
    ('12/03/2026','2026-03-12', 56545.49,'TCE-TO — Tribunal de Contas',        '202600000000218','mar/26','TED — GOVERNO DO EST 01.786.029/0001-03'),
    ('16/03/2026','2026-03-16',  1637.13,'A identificar — Município de Palmas','—',              'mar/26','Ordem Bancária — MUNICIPIO DE PALMAS ⚠ sem NF'),
    ('16/03/2026','2026-03-16',  2182.83,'Município de Palmas',                '202600000000301','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  2182.84,'Município de Palmas',                '202600000000296','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  3429.28,'Ministério Público do TO (MPTO)',    '202600000000349','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  4502.10,'Município de Palmas',                '202600000000297','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  4502.11,'Município de Palmas',                '202600000000310','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  6139.22,'Ministério Público do TO (MPTO)',    '202600000000173','fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  6139.23,'Ministério Público do TO (MPTO)',    '202600000000173','fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  8185.63,'Município de Palmas (ARCAF)',         '202600000000307','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  8185.64,'Município de Palmas (ARCAF)',         '202600000000307','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16',  9563.24,'Ministério Público do TO (MPTO)',    '202600000000354','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16', 12859.84,'Município de Palmas',                '202600000000083','jan/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16', 18008.39,'SEDUC-TO',                           '202600000000088','jan/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16', 19645.52,'PREVIPALMAS — Prev. Social Palmas',  '202600000000341','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16', 24556.90,'Município de Palmas',                '202600000000299','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('16/03/2026','2026-03-16', 28474.95,'CBMTO — Corpo de Bombeiros',         '202600000000217','mar/26','TED — GOVERNO DO EST 01.786.029/0001-03'),
    ('16/03/2026','2026-03-16', 35862.16,'CBMTO — Corpo de Bombeiros',         '202600000000217','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('17/03/2026','2026-03-17',  1637.13,'A identificar — Município de Palmas','—',              'mar/26','Ordem Bancária — MUNICIPIO DE PALMAS ⚠ sem NF'),
    ('17/03/2026','2026-03-17', 36016.79,'CBMTO — Corpo de Bombeiros',         '202600000000217','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('17/03/2026','2026-03-17', 49113.80,'Município de Palmas',                '202600000000187','fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('17/03/2026','2026-03-17', 49113.81,'Município de Palmas',                '202600000000187','fev/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('18/03/2026','2026-03-18', 19038.27,'PREVIPALMAS — Prev. Social Palmas',  '202600000000341','mar/26','Ordem Bancária — ORDENS BANCARIAS'),
    ('18/03/2026','2026-03-18',193359.33,'UNITINS — Univ. Estadual do TO',     '202600000000430','mar/26','TED — GOVERNO DO EST 01.786.029/0001-03'),
    ('19/03/2026','2026-03-19', 24556.90,'Município de Palmas',                '202600000000299','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('19/03/2026','2026-03-19', 35862.16,'CBMTO — Corpo de Bombeiros',         '202600000000217','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('19/03/2026','2026-03-19', 73670.71,'UFT — Fundação Univ. Federal do TO', '202600000000441','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('26/03/2026','2026-03-26',  9667.72,'Município de Palmas',                '202600000000245','mar/26','Ordem Bancária — MUNICIPIO DE PALMAS'),
    ('26/03/2026','2026-03-26', 12859.84,'Município de Palmas',                '202600000000083','jan/26','Ordem Bancária — ORDENS BANCARIAS'),
    ('26/03/2026','2026-03-26', 24556.90,'Município de Palmas',                '202600000000299','mar/26','Ordem Bancária — ORDENS BANCARIAS'),
    ('26/03/2026','2026-03-26', 49113.81,'Município de Palmas',                '202600000000187','fev/26','Ordem Bancária — ORDENS BANCARIAS'),
]

# 1 crédito EXCLUÍDO — NF dez/2025, já tributada sob competência
EXCLUIDOS = [
    ('19/03/2026','2026-03-19', 26194.03,
     'UFT — Fund. Univ. Federal do TO',
     '202500000001384', 'dez/2025',
     'Ordem Bancária — MUNICIPIO DE PALMAS',
     'NF emitida em dez/2025 — já tributada sob regime de competência 2025'),
]

# Não tributa
BB_RENDE = [
    ('BB Rende Fácil / Aplicação Automática', 10, 1802218.07),
]
BB_INTERNO = [
    ('Transferências Internas — Grupo Montana', 5, 415000.00),
]
BB_PENDENTE = [
    ('BB PENDENTE sem NF identificada', 2, 883.52),
]
BRB_INV = [
    ('BRB — Resgates CDB/RDB Mar/2026', 3, 617500.00),
    ('BRB — Resgates FI BRB Federal Invest', 2, 140000.00),
    ('BRB — Créditos Juros CDB Automático', 2, 25.33),
    ('BRB — Resgate CDB/RDB 23/03', 1, 523785.55),
    ('BRB — Resgate CDB/RDB 24/03', 1, 50000.01),
    ('BRB — Resgate CDB/RDB 26/03', 1, 300000.00),
]

# Diferido — NFs 2026 emitidas e não recebidas até 31/03
DIFERIDO = [
    ('jan/2026',  84, 'MPTO, Palmas, SEDUC-TO e outros',  700000.00),
    ('fev/2026', 122, 'Palmas, UNITINS, DETRAN-TO e outros', 3500000.00),
    ('mar/2026', 151, 'Palmas, FCP, CBMTO, TCE-TO, UFT e outros', 7200000.00),
]

# ─── CÁLCULOS ──────────────────────────────────────────────────────────────
BASE = sum(v for _,_,v,*_ in TRIBUTAVEIS)
EXCLUIDO_TOTAL = sum(v for _,_,v,*_ in EXCLUIDOS)
PIS  = BASE * 0.0065
COF  = BASE * 0.030
TOT  = BASE * 0.0365

TOTAL_ENTRADA = BASE + EXCLUIDO_TOTAL  # = 1.521.746,81

# ─── WORKBOOK ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════
# ABA 1 — RESUMO EXECUTIVO
# ══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Resumo Executivo'
ws.sheet_view.showGridLines = False
ws.row_dimensions[1].height = 14
ws.row_dimensions[2].height = 32
ws.row_dimensions[3].height = 16

for col in range(1, 9):
    ws.column_dimensions[get_column_letter(col)].width = [3, 32, 14, 14, 12, 14, 14, 3][col-1]

# Título
merge(ws,2,2,2,7)
c = ws.cell(2,2,'APURAÇÃO PIS/COFINS — MARÇO/2026')
c.font = Font(bold=True, color='FFFFFF', size=15)
c.alignment = Alignment(horizontal='center', vertical='center')
c.fill = fill(C_VERDE_ESC)
c.border = border_med()

merge(ws,3,2,3,7)
sub = f'Montana Segurança Privada Ltda  ·  Lucro Real  ·  PIS/COFINS Cumulativo 3,65%  ·  Regime de Caixa (desde jan/2026)  ·  {datetime.now():%d/%m/%Y}'
c2 = ws.cell(3,2,sub)
c2.font = Font(bold=False, color='FFFFFF', size=9)
c2.alignment = Alignment(horizontal='center', vertical='center')
c2.fill = fill(C_VERDE_MED)
c2.border = border_thin()

# ALERTA TRANSIÇÃO
ws.row_dimensions[5].height = 28
merge(ws,5,2,5,7)
alerta = ws.cell(5,2,'⚠  REGIME DE TRANSIÇÃO: empresa apurava por COMPETÊNCIA até dez/2025. A partir de jan/2026 → CAIXA.')
alerta.font = Font(bold=True, color='7B3F00', size=10)
alerta.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
alerta.fill = fill(C_LARANJA)
alerta.border = border_med()

# Seção COMPOSIÇÃO
row = 7
header_cell(ws,row,2,'COMPOSIÇÃO DOS CRÉDITOS BANCÁRIOS — BB + BRB',C_VERDE_MED,align='left')
header_cell(ws,row,3,'Qtd',C_VERDE_MED)
header_cell(ws,row,4,'Valor (R$)',C_VERDE_MED)
header_cell(ws,row,5,'Base PIS/COFINS',C_VERDE_MED)
merge(ws,row,2,row,2)
ws.row_dimensions[row].height = 18

composicao = [
    ('BB — Créditos Tributáveis (NFs de 2026 — CAIXA)',     len(TRIBUTAVEIS),  BASE,           '✅ SIM'),
    ('BB — Excluído — NF dez/2025 (já tributada 2025)',     len(EXCLUIDOS),    EXCLUIDO_TOTAL, '⛔ NÃO — exercício anterior'),
    ('BB — Rende Fácil (INVESTIMENTO)',                     10,                1802218.07,     '❌ NÃO'),
    ('BB — Transferências Internas (INTERNO Montana)',       5,                 415000.00,      '❌ NÃO'),
    ('BB — PENDENTE sem NF identificada',                   2,                 883.52,         '❌ NÃO'),
    ('BRB — Resgates CDB/FI BRB (INVESTIMENTO)',            11,                1478810.89,     '❌ NÃO'),
]
total_geral = sum(v for _,_,v,_ in composicao)
total_qtd   = sum(q for _,q,_,_ in composicao)

for i,(desc,qtd,val,base_flag) in enumerate(composicao):
    r = row+1+i
    ws.row_dimensions[r].height = 15
    bg = C_VERDE_ROWS if i%2==0 else None
    if 'Tributáveis' in desc: bg = C_VERDE_CLAR
    if 'Excluído' in desc:    bg = C_EXCLUIDO
    data_cell(ws,r,2,desc,bg=bg)
    data_cell(ws,r,3,qtd,align='center',bg=bg)
    data_cell(ws,r,4,val,money_fmt(),bg=bg,align='right')
    data_cell(ws,r,5,base_flag,align='center',bg=bg,
              bold='✅' in base_flag, color=(C_VERDE_ESC if '✅' in base_flag else ('C62828' if '⛔' in base_flag else '616161')))

tot_r = row+1+len(composicao)
ws.row_dimensions[tot_r].height = 16
header_cell(ws,tot_r,2,'TOTAL GERAL',C_VERDE_ESC,align='left')
header_cell(ws,tot_r,3,total_qtd,C_VERDE_ESC)
data_cell(ws,tot_r,4,total_geral,money_fmt(),bg=C_VERDE_ESC,bold=True,color='FFFFFF',align='right')
ws.cell(tot_r,4).font = Font(bold=True,color='FFFFFF',size=10)
data_cell(ws,tot_r,5,'',bg=C_VERDE_ESC)

# Seção APURAÇÃO
row2 = tot_r + 2
ws.row_dimensions[row2].height = 18
header_cell(ws,row2,2,'APURAÇÃO DO IMPOSTO',C_VERDE_MED,align='left')
header_cell(ws,row2,3,'',C_VERDE_MED)
header_cell(ws,row2,4,'Valor (R$)',C_VERDE_MED)
merge(ws,row2,2,row2,2)

apuracao = [
    ('BASE TRIBUTÁVEL (48 créditos — NFs 2026)',        BASE,   True,  C_VERDE_CLAR),
    ('(-) Excluído NF dez/2025 — já tributada 2025',   -EXCLUIDO_TOTAL, False, C_EXCLUIDO),
    ('Base de cálculo efetiva',                         BASE,   True,  C_VERDE_CLAR),
    ('PIS — 0,65% (Regime Cumulativo)',                 PIS,    False, None),
    ('COFINS — 3,00% (Regime Cumulativo)',              COF,    False, None),
    ('TOTAL PIS + COFINS  —  3,65% cumulativo',        TOT,    True,  C_VERDE_CLAR),
]

for i,(desc,val,bold,bg) in enumerate(apuracao):
    r = row2+1+i
    ws.row_dimensions[r].height = 15
    row_bg = bg or (C_CINZA_CLAR if i%2==0 else None)
    data_cell(ws,r,2,desc,bg=row_bg,bold=bold)
    if val is not None:
        data_cell(ws,r,4,val,money_fmt(),bg=row_bg,bold=bold,align='right')
    else:
        data_cell(ws,r,4,'—',bg=row_bg,align='center')
    data_cell(ws,r,3,'',bg=row_bg)

# DARF
darf_r = row2+1+len(apuracao)+1
ws.row_dimensions[darf_r].height = 20
merge(ws,darf_r,2,darf_r,5)
darf_c = ws.cell(darf_r,2,
    f'⚠  DARF — Vencimento: 27/04/2026  ·  PIS cód. 6912 (R${PIS:,.2f})  ·  COFINS cód. 5856 (R${COF:,.2f})  ·  Regime Cumulativo — sem créditos de entrada a abater')
darf_c.font = Font(bold=True, color='B71C1C', size=10)
darf_c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
darf_c.fill = fill('FFCDD2')
darf_c.border = border_med()

# Nota transição
nota_r = darf_r+2
ws.row_dimensions[nota_r].height = 28
merge(ws,nota_r,2,nota_r,7)
nota = ws.cell(nota_r,2,
    'NOTA FISCAL EXCLUÍDA: R$ 26.194,03 (UFT — NF 202500000001384, competência dez/2025) foi paga em mar/2026 '
    'mas JÁ FOI TRIBUTADA sob regime de competência em 2025. Excluída da base de caixa 2026 para evitar bitributação. '
    'Documentação de suporte: NF WebISS + DARF dez/2025.')
nota.font = Font(italic=True, color='5D4037', size=9)
nota.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
nota.fill = fill(C_EXCLUIDO)
nota.border = border_thin()

# ══════════════════════════════════════════════════════════════════════════
# ABA 2 — CRÉDITOS TRIBUTÁVEIS
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Créditos Tributáveis')
ws2.sheet_view.showGridLines = False

cols2 = [3,12,8,32,24,10,30,12]
for i,w in enumerate(cols2): ws2.column_dimensions[get_column_letter(i+1)].width = w

# Título
merge(ws2,1,1,1,8)
c = ws2.cell(1,1,'CRÉDITOS TRIBUTÁVEIS — PIS/COFINS MARÇO/2026  |  Montana Segurança Privada Ltda  |  Regime de Caixa')
c.font = Font(bold=True, color='FFFFFF', size=11)
c.alignment = Alignment(horizontal='center', vertical='center')
c.fill = fill(C_VERDE_ESC)
c.border = border_med()
ws2.row_dimensions[1].height = 22

merge(ws2,2,1,2,8)
c2 = ws2.cell(2,1,'Apenas créditos correspondentes a NFs emitidas em 2026 — base do regime de caixa vigente desde jan/2026')
c2.font = Font(italic=True, color='FFFFFF', size=9)
c2.alignment = Alignment(horizontal='center', vertical='center')
c2.fill = fill(C_VERDE_MED)
c2.border = border_thin()
ws2.row_dimensions[2].height = 15

hdrs2 = ['#','Data','Valor (R$)','Tomador','NF / Referência','Comp. NF','Histórico Banco','PIS+COFINS (3,65%)']
for j,h in enumerate(hdrs2):
    header_cell(ws2,3,j+1,h,C_VERDE_MED)
ws2.row_dimensions[3].height = 18

for i,(data_br,_,val,tomador,nf,comp,hist) in enumerate(TRIBUTAVEIS):
    r = 4+i
    ws2.row_dimensions[r].height = 14
    bg = C_VERDE_ROWS if i%2==0 else None
    flag = '⚠' in hist  # uncertain entries
    if flag: bg = C_AMARELO
    data_cell(ws2,r,1,i+1,align='center',bg=bg)
    data_cell(ws2,r,2,data_br,align='center',bg=bg)
    data_cell(ws2,r,3,val,money_fmt(),bg=bg,align='right',bold=flag)
    data_cell(ws2,r,4,tomador,bg=bg)
    data_cell(ws2,r,5,nf,align='center',bg=bg)
    data_cell(ws2,r,6,comp,align='center',bg=bg)
    data_cell(ws2,r,7,hist[:55],bg=bg)
    data_cell(ws2,r,8,val*0.0365,money_fmt(),bg=bg,align='right')

tot2 = 4+len(TRIBUTAVEIS)
ws2.row_dimensions[tot2].height = 16
for j in range(1,9):
    header_cell(ws2,tot2,j,'',C_VERDE_ESC)
ws2.cell(tot2,1,'TOTAL').font = Font(bold=True,color='FFFFFF',size=10)
ws2.cell(tot2,1).fill = fill(C_VERDE_ESC)
ws2.cell(tot2,1).alignment = Alignment(horizontal='center',vertical='center')
ws2.cell(tot2,3,BASE).number_format = money_fmt()
ws2.cell(tot2,3).font = Font(bold=True,color='FFFFFF',size=10)
ws2.cell(tot2,3).fill = fill(C_VERDE_ESC)
ws2.cell(tot2,3).alignment = Alignment(horizontal='right',vertical='center')
ws2.cell(tot2,8,BASE*0.0365).number_format = money_fmt()
ws2.cell(tot2,8).font = Font(bold=True,color='FFFFFF',size=10)
ws2.cell(tot2,8).fill = fill(C_VERDE_ESC)
ws2.cell(tot2,8).alignment = Alignment(horizontal='right',vertical='center')

# Legenda
leg_r = tot2+2
merge(ws2,leg_r,1,leg_r,8)
leg = ws2.cell(leg_r,1,'⚠ Linhas em amarelo: NF não encontrada individualmente no ERP — vinculação estimada por valor/tomador. '
               'Contador deve confirmar a NF exata no WebISS antes de emitir o DARF.')
leg.font = Font(italic=True, color='7B3F00', size=9)
leg.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
leg.fill = fill(C_AMARELO)
leg.border = border_thin()
ws2.row_dimensions[leg_r].height = 22

# ══════════════════════════════════════════════════════════════════════════
# ABA 3 — EXERCÍCIOS ANTERIORES (EXCLUÍDOS)
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Excluídos — Exerc. Anterior')
ws3.sheet_view.showGridLines = False

cols3 = [3,12,10,32,26,10,30,30]
for i,w in enumerate(cols3): ws3.column_dimensions[get_column_letter(i+1)].width = w

merge(ws3,1,1,1,8)
c = ws3.cell(1,1,'CRÉDITOS EXCLUÍDOS DA BASE — NFs DE 2024/2025 JÁ TRIBUTADAS SOB COMPETÊNCIA')
c.font = Font(bold=True, color='FFFFFF', size=11)
c.alignment = Alignment(horizontal='center', vertical='center')
c.fill = fill('C62828')
c.border = border_med()
ws3.row_dimensions[1].height = 22

merge(ws3,2,1,2,8)
c2 = ws3.cell(2,1,
    'A empresa apurava PIS/COFINS por competência até dez/2025. Pagamentos recebidos em 2026 referentes '
    'a NFs de 2024/2025 NÃO integram a base de caixa 2026 — já foram tributados no exercício de emissão.')
c2.font = Font(italic=True, color='FFFFFF', size=9)
c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
c2.fill = fill('E53935')
c2.border = border_thin()
ws3.row_dimensions[2].height = 20

hdrs3 = ['#','Data Recebto.','Valor (R$)','Tomador','NF','Comp. NF','Histórico Banco','Motivo da Exclusão']
for j,h in enumerate(hdrs3):
    c = ws3.cell(3,j+1,h)
    c.font = Font(bold=True,color='FFFFFF',size=10)
    c.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
    c.fill = fill('E53935')
    c.border = border_thin()
ws3.row_dimensions[3].height = 18

for i,(data_br,_,val,tom,nf,comp,hist,motivo) in enumerate(EXCLUIDOS):
    r = 4+i
    ws3.row_dimensions[r].height = 14
    bg = C_EXCLUIDO
    data_cell(ws3,r,1,i+1,align='center',bg=bg)
    data_cell(ws3,r,2,data_br,align='center',bg=bg)
    data_cell(ws3,r,3,val,money_fmt(),bg=bg,align='right')
    data_cell(ws3,r,4,tom,bg=bg)
    data_cell(ws3,r,5,nf,align='center',bg=bg)
    data_cell(ws3,r,6,comp,align='center',bg=bg,color='C62828',bold=True)
    data_cell(ws3,r,7,hist[:55],bg=bg)
    data_cell(ws3,r,8,motivo,bg=bg,color='C62828')

tot3 = 4+len(EXCLUIDOS)
ws3.row_dimensions[tot3].height = 16
for j in range(1,9):
    c = ws3.cell(tot3,j)
    c.fill = fill('C62828')
    c.border = border_thin()
ws3.cell(tot3,1,'TOTAL').font = Font(bold=True,color='FFFFFF',size=10)
ws3.cell(tot3,1).alignment = Alignment(horizontal='center',vertical='center')
ws3.cell(tot3,3,EXCLUIDO_TOTAL).number_format = money_fmt()
ws3.cell(tot3,3).font = Font(bold=True,color='FFFFFF',size=10)
ws3.cell(tot3,3).alignment = Alignment(horizontal='right',vertical='center')

# nota
nota3_r = tot3+2
merge(ws3,nota3_r,1,nota3_r,8)
nota3 = ws3.cell(nota3_r,1,
    'DOCUMENTAÇÃO OBRIGATÓRIA: Manter cópia da NF 202500000001384 (WebISS) + DARF de dez/2025 em que '
    'esta NF foi tributada. Em caso de fiscalização, comprovar que não houve omissão de receita — '
    'apenas aplicação da regra de transição competência → caixa.')
nota3.font = Font(italic=True, color='5D4037', size=9)
nota3.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
nota3.fill = fill(C_EXCLUIDO)
nota3.border = border_thin()
ws3.row_dimensions[nota3_r].height = 28

# ══════════════════════════════════════════════════════════════════════════
# ABA 4 — NÃO TRIBUTA
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Não Tributa')
ws4.sheet_view.showGridLines = False
for i,w in enumerate([3,40,8,14,30,3]): ws4.column_dimensions[get_column_letter(i+1)].width = w

merge(ws4,1,1,1,5)
c = ws4.cell(1,1,'CRÉDITOS NÃO TRIBUTÁVEIS — BB + BRB  |  Março/2026')
c.font = Font(bold=True,color='FFFFFF',size=11)
c.alignment = Alignment(horizontal='center',vertical='center')
c.fill = fill('546E7A')
c.border = border_med()
ws4.row_dimensions[1].height = 22

for j,h in enumerate(['#','Descrição','Qtd','Valor (R$)','Motivo / Classificação']):
    header_cell(ws4,2,j+1,h,'546E7A')
ws4.row_dimensions[2].height = 18

nao_trib = (
    [(d,q,v,'INVESTIMENTO — Resgate de aplicação não é receita de serviço') for d,q,v in BB_RENDE] +
    [(d,q,v,'INTERNO — Transferência entre contas do Grupo Montana') for d,q,v in BB_INTERNO] +
    [(d,q,v,'PENDENTE — Créditos sem NF identificada') for d,q,v in BB_PENDENTE] +
    [(d,q,v,'INVESTIMENTO — Resgates CDB/RDB/FI BRB (conta exclusiva de investimentos)') for d,q,v in BRB_INV]
)
cores_nt = ['B0BEC5','CFD8DC','ECEFF1','B0BEC5','CFD8DC','ECEFF1','B0BEC5','CFD8DC','ECEFF1','B0BEC5']
for i,(desc,qtd,val,motivo) in enumerate(nao_trib):
    r = 3+i
    bg = cores_nt[i%len(cores_nt)]
    ws4.row_dimensions[r].height = 14
    data_cell(ws4,r,1,i+1,align='center',bg=bg)
    data_cell(ws4,r,2,desc,bg=bg)
    data_cell(ws4,r,3,qtd,align='center',bg=bg)
    data_cell(ws4,r,4,val,money_fmt(),bg=bg,align='right')
    data_cell(ws4,r,5,motivo,bg=bg)

tot4 = 3+len(nao_trib)
ws4.row_dimensions[tot4].height = 16
for j in range(1,6):
    c = ws4.cell(tot4,j)
    c.fill = fill('546E7A')
    c.border = border_thin()
ws4.cell(tot4,2,'TOTAL NÃO TRIBUTA').font = Font(bold=True,color='FFFFFF',size=10)
ws4.cell(tot4,2).alignment = Alignment(horizontal='left',vertical='center')
ws4.cell(tot4,3,sum(q for _,q,_,_ in nao_trib)).font = Font(bold=True,color='FFFFFF',size=10)
ws4.cell(tot4,3).alignment = Alignment(horizontal='center',vertical='center')
ws4.cell(tot4,4,sum(v for _,_,v,_ in nao_trib)).number_format = money_fmt()
ws4.cell(tot4,4).font = Font(bold=True,color='FFFFFF',size=10)
ws4.cell(tot4,4).alignment = Alignment(horizontal='right',vertical='center')

nota4_r = tot4+2
merge(ws4,nota4_r,1,nota4_r,5)
nota4 = ws4.cell(nota4_r,1,
    'NOTA BRB: Conta BRB (Ag. 031 — Cta. 031.015.474-0) é exclusiva para gestão financeira (CDB, RDB, FI). '
    'Nenhum pagamento de cliente transita pelo BRB. Todos os R$ 1.478.810,89 de créditos BRB são INVESTIMENTO.')
nota4.font = Font(italic=True, size=9, color='37474F')
nota4.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
nota4.fill = fill('ECEFF1')
nota4.border = border_thin()
ws4.row_dimensions[nota4_r].height = 22

# ══════════════════════════════════════════════════════════════════════════
# ABA 5 — DIFERIDO
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Diferido — NFs Não Recebidas')
ws5.sheet_view.showGridLines = False
for i,w in enumerate([3,12,8,40,18,3]): ws5.column_dimensions[get_column_letter(i+1)].width = w

merge(ws5,1,1,1,5)
c = ws5.cell(1,1,'DIFERIDO — NFs 2026 EMITIDAS E NÃO RECEBIDAS ATÉ 31/03/2026  |  Montana Segurança')
c.font = Font(bold=True,color='FFFFFF',size=11)
c.alignment = Alignment(horizontal='center',vertical='center')
c.fill = fill('F57F17')
c.border = border_med()
ws5.row_dimensions[1].height = 22

merge(ws5,2,1,2,5)
c2 = ws5.cell(2,1,'Regime de caixa: estas NFs tributarão no mês em que o dinheiro for efetivamente recebido. Não entram na base de março/2026.')
c2.font = Font(italic=True,color='FFFFFF',size=9)
c2.alignment = Alignment(horizontal='center',vertical='center')
c2.fill = fill('F9A825')
c2.border = border_thin()
ws5.row_dimensions[2].height = 15

for j,h in enumerate(['#','Competência','NFs','Principais Tomadores','Valor Estimado (R$)']):
    header_cell(ws5,3,j+1,h,'F57F17')
ws5.row_dimensions[3].height = 18

cores_dif = ['FFF9C4','FFFDE7','FFF8E1']
for i,(comp,qtd,toms,val) in enumerate(DIFERIDO):
    r = 4+i
    bg = cores_dif[i%3]
    ws5.row_dimensions[r].height = 16
    data_cell(ws5,r,1,i+1,align='center',bg=bg)
    data_cell(ws5,r,2,comp,align='center',bg=bg,bold=True)
    data_cell(ws5,r,3,qtd,align='center',bg=bg)
    data_cell(ws5,r,4,toms,bg=bg)
    data_cell(ws5,r,5,val,money_fmt(),bg=bg,align='right')

tot5 = 4+len(DIFERIDO)
ws5.row_dimensions[tot5].height = 16
for j in range(1,6):
    c = ws5.cell(tot5,j)
    c.fill = fill('F57F17')
    c.border = border_thin()
ws5.cell(tot5,2,'TOTAL').font = Font(bold=True,color='FFFFFF',size=10)
ws5.cell(tot5,2).alignment = Alignment(horizontal='center',vertical='center')
ws5.cell(tot5,3,sum(q for _,q,_,_ in DIFERIDO)).font = Font(bold=True,color='FFFFFF',size=10)
ws5.cell(tot5,3).alignment = Alignment(horizontal='center',vertical='center')
ws5.cell(tot5,5,sum(v for _,_,_,v in DIFERIDO)).number_format = money_fmt()
ws5.cell(tot5,5).font = Font(bold=True,color='FFFFFF',size=10)
ws5.cell(tot5,5).alignment = Alignment(horizontal='right',vertical='center')

nota5_r = tot5+2
merge(ws5,nota5_r,1,nota5_r,5)
nota5 = ws5.cell(nota5_r,1,
    'BACKLOG ADICIONAL 2024/2025: A empresa possui ainda 2.408 NFs emitidas em 2024/2025 com pagamento pendente '
    '(R$ 74.589.273,47). Quando recebidas em 2026 ou anos seguintes, NÃO tributam PIS/COFINS novamente — '
    'já foram tributadas sob regime de competência nos respectivos exercícios.')
nota5.font = Font(italic=True, size=9, color='5D4037')
nota5.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
nota5.fill = fill(C_LARANJA)
nota5.border = border_thin()
ws5.row_dimensions[nota5_r].height = 28

# ─── SALVAR ────────────────────────────────────────────────────────────────
path = f'Apuracao_PISCOFINS_Seguranca_202603_FINAL.xlsx'
wb.save(path)
print(f'Salvo: {path}')
print(f'Base tributável: R$ {BASE:,.2f}')
print(f'Excluído (dez/25): R$ {EXCLUIDO_TOTAL:,.2f}')
print(f'PIS (0,65%):   R$ {PIS:,.2f}')
print(f'COFINS (3,00%): R$ {COF:,.2f}')
print(f'TOTAL:         R$ {TOT:,.2f}')
print(f'Vencimento DARF: 27/04/2026')
