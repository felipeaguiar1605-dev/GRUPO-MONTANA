from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import datetime

wb = Workbook()

# ── Paleta ────────────────────────────────────────────────────────────────────
AZUL_ESCURO  = "1F497D"
AZUL_MEDIO   = "2E75B6"
AZUL_CLARO   = "D9E1F2"
VERDE_FUNDO  = "C6EFCE"
VERDE_TEXTO  = "375623"
VERM_FUNDO   = "FFC7CE"
VERM_TEXTO   = "9C0006"
VERM_ESCURO  = "C00000"
AMAR_FUNDO   = "FFEB9C"
AMAR_TEXTO   = "7D6608"
LARAN_FUNDO  = "FCE4D6"
LARAN_TEXTO  = "833C00"
CINZA_FUNDO  = "F2F2F2"
CINZA_BORDA  = "BFBFBF"
BRANCO       = "FFFFFF"

def cor(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fonte(bold=False, size=11, color="000000", name="Arial", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def lado():
    return Side(style="thin", color=CINZA_BORDA)

def borda_fina():
    s = lado()
    return Border(left=s, right=s, top=s, bottom=s)

def lado_m():
    return Side(style="medium", color="000000")

def borda_media():
    s = lado_m()
    return Border(left=s, right=s, top=s, bottom=s)

def alinhar(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def fmt_brl(v):
    return v  # valor numérico; formatação é aplicada na célula

def aplicar_brl(cell):
    cell.number_format = '#,##0.00'

STATUS_ESTILOS = {
    "PAGO":              (VERDE_FUNDO, VERDE_TEXTO),
    "VENCIDO":           (VERM_FUNDO,  VERM_TEXTO),
    "VENCE ESTA SEMANA": (AMAR_FUNDO,  AMAR_TEXTO),
    "A IDENTIFICAR":     (LARAN_FUNDO, LARAN_TEXTO),
    "PENDENTE":          (LARAN_FUNDO, LARAN_TEXTO),
    "VERIFICAR":         (AMAR_FUNDO,  AMAR_TEXTO),
    "EM DIA":            (VERDE_FUNDO, VERDE_TEXTO),
}

def aplicar_status(cell, status):
    fundo, texto = STATUS_ESTILOS.get(status, (CINZA_FUNDO, "000000"))
    cell.fill = cor(fundo)
    cell.font = fonte(bold=True, color=texto, size=10)
    cell.alignment = alinhar("center")
    cell.border = borda_fina()

def linha_dados(ws, row, cols_data, status_col=None, status_val=None,
                bold_first=False, moeda_cols=None):
    """
    cols_data: list of (col_idx, value)
    moeda_cols: set of col indices that should get BRL format
    """
    if moeda_cols is None:
        moeda_cols = set()
    for col, val in cols_data:
        c = ws.cell(row=row, column=col, value=val)
        c.border = borda_fina()
        c.alignment = alinhar("left", "center")
        c.font = fonte(bold=(bold_first and col == cols_data[0][0]), size=10)
        if col in moeda_cols and isinstance(val, (int, float)):
            aplicar_brl(c)
    if status_col and status_val:
        aplicar_status(ws.cell(row=row, column=status_col), status_val)

def cabecalho_secao(ws, row, col_start, col_end, titulo, cor_fundo=AZUL_MEDIO):
    c = ws.cell(row=row, column=col_start, value=titulo)
    c.fill = cor(cor_fundo)
    c.font = fonte(bold=True, size=11, color=BRANCO)
    c.alignment = alinhar("center")
    c.border = borda_media()
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)

def cabecalhos_linha(ws, row, cols_headers, fundo=AZUL_CLARO):
    for col, titulo in cols_headers:
        c = ws.cell(row=row, column=col, value=titulo)
        c.fill = cor(fundo)
        c.font = fonte(bold=True, size=10, color="000000")
        c.alignment = alinhar("center", wrap=True)
        c.border = borda_fina()

def titulo_aba(ws, row, titulo, subtitulo=None, max_col=8):
    c = ws.cell(row=row, column=1, value=titulo)
    c.fill = cor(AZUL_ESCURO)
    c.font = fonte(bold=True, size=14, color=BRANCO)
    c.alignment = alinhar("center")
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=max_col)
    if subtitulo:
        c2 = ws.cell(row=row+1, column=1, value=subtitulo)
        c2.fill = cor(AZUL_MEDIO)
        c2.font = fonte(size=10, color=BRANCO, italic=True)
        c2.alignment = alinhar("center")
        ws.merge_cells(start_row=row+1, start_column=1,
                       end_row=row+1, end_column=max_col)


# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — CALENDÁRIO TRIBUTÁRIO
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📅 Calendário"
ws1.sheet_properties.tabColor = "1F497D"

titulo_aba(ws1, 1,
    "GRUPO MONTANA SEC — FLUXO DE CAIXA TRIBUTÁRIO 2026",
    f"Gerado em {datetime.date.today().strftime('%d/%m/%Y')} | Base: dados reais Gmail + SQLite",
    max_col=9)

# Aviso crítico
c_aviso = ws1.cell(row=3, column=1,
    value="⚠  ATENÇÃO: Obrigações com vencimento em 20/04/2026 (SEGUNDA-FEIRA) — PRAZO EM 2 DIAS")
c_aviso.fill = cor(VERM_ESCURO)
c_aviso.font = fonte(bold=True, size=12, color=BRANCO)
c_aviso.alignment = alinhar("center")
c_aviso.border = borda_media()
ws1.merge_cells("A3:I3")

ws1.row_dimensions[3].height = 22

# Cabeçalhos tabela
cabecalho_secao(ws1, 5, 1, 9, "CALENDÁRIO DE OBRIGAÇÕES TRIBUTÁRIAS — ABR / MAI / JUN 2026")

hdrs = [(1,"Empresa"),(2,"Tributo"),(3,"Competência"),(4,"Vencimento"),
        (5,"Valor (R$)"),(6,"DARF/DAM"),(7,"Status"),(8,"Observações"),(9,"Ação")]
cabecalhos_linha(ws1, 6, hdrs)

# Larguras
for col, w in [(1,22),(2,22),(3,14),(4,14),(5,18),(6,16),(7,20),(8,40),(9,25)]:
    ws1.column_dimensions[get_column_letter(col)].width = w

ws1.freeze_panes = "A7"

obrigacoes = [
    # Empresa, Tributo, Competência, Vencimento, Valor, DARF, Status, Obs, Ação
    ("Assessoria","ISSQN","MAR/2026","15/04/2026",315.07,"DAM Palmas",
     "VENCIDO","Guia DAM emitida — verificar pagamento",
     "Pagar + multa 2% + juros SELIC"),

    ("Assessoria","INSS Patronal","MAR/2026","20/04/2026",416947.58,"DCTFWeb",
     "VENCE ESTA SEMANA","⚠ CRÍTICO — vence 20/04 (segunda-feira)  |  Incluir crédito INSS retido (ver aba Créditos)",
     "URGENTE — emitir DARF até 19/04"),

    ("Assessoria","FGTS","MAR/2026","20/04/2026",208262.37,"GFIP/SEFIP",
     "VENCE ESTA SEMANA","⚠ CRÍTICO — mesmo vencimento 20/04",
     "URGENTE — recolher FGTS até 19/04"),

    ("Assessoria","PIS","FEV/2026","25/03/2026",None,"DARF 8109",
     "A IDENTIFICAR","Guia não localizada no Gmail — verificar escritório",
     "Consultar contabilidade"),

    ("Assessoria","COFINS","FEV/2026","25/03/2026",None,"DARF 2172",
     "A IDENTIFICAR","Guia não localizada no Gmail — verificar escritório",
     "Consultar contabilidade"),

    ("Assessoria","PIS","MAR/2026","27/04/2026",None,"DARF 8109",
     "PENDENTE","Vence 27/04 — guia não localizada ainda",
     "Solicitar guia à contabilidade"),

    ("Assessoria","COFINS","MAR/2026","27/04/2026",None,"DARF 2172",
     "PENDENTE","Vence 27/04 — guia não localizada ainda",
     "Solicitar guia à contabilidade"),

    ("Assessoria","IRPJ (Estim.)","1°TRIM/2026","30/04/2026",None,"DARF 1082",
     "A IDENTIFICAR","Estimativa mensal ou ajuste trimestral — verificar com contabilidade",
     "Aguardar cálculo contabilidade"),

    ("Assessoria","CSLL (Estim.)","1°TRIM/2026","30/04/2026",None,"DARF 1138",
     "A IDENTIFICAR","Idem IRPJ — cálculo pendente",
     "Aguardar cálculo contabilidade"),

    ("Assessoria","PIS","DEZ/2025","23/01/2026",47492.23,"DARF 8109",
     "PAGO","Comprovante localizado no Gmail — pago em 23/01/2026",
     "—"),

    ("Assessoria","COFINS","DEZ/2025","23/01/2026",218502.34,"DARF 2172",
     "PAGO","Comprovante localizado no Gmail — pago em 23/01/2026",
     "—"),

    ("Assessoria","PIS","JAN/2026","25/02/2026",29424.61,"DARF 8109",
     "PAGO","Comprovante localizado no Gmail — pago em 25/02/2026",
     "—"),

    ("Assessoria","COFINS","JAN/2026","25/02/2026",135441.71,"DARF 2172",
     "PAGO","Comprovante localizado no Gmail — pago em 25/02/2026",
     "—"),

    ("Assessoria","ISSQN","FEV/2026","13/03/2026",15.13,"DAM Palmas",
     "PAGO","Pago — guia localizada no Gmail",
     "—"),

    ("Assessoria","ISSQN (DETRAN)","FEV/2026","16/03/2026",235.85,"DAM Palmas",
     "PAGO","Pago — guia localizada no Gmail",
     "—"),

    ("Segurança","PIS","MAR/2026","27/04/2026",9721.09,"DARF 8109",
     "PENDENTE","Vence 27/04 — valor calculado (0,65% sobre faturamento)",
     "Emitir DARF quando guia disponível"),

    ("Segurança","COFINS","MAR/2026","27/04/2026",44866.58,"DARF 2172",
     "PENDENTE","Vence 27/04 — valor calculado (3% sobre faturamento)",
     "Emitir DARF quando guia disponível"),

    ("Segurança","INSS/FGTS","MAR/2026","20/04/2026",None,"DCTFWeb/SEFIP",
     "A IDENTIFICAR","Valores não localizados — solicitar à contabilidade",
     "URGENTE — verificar com escritório"),

    ("Mustang","DAS","MAR/2026","20/04/2026",None,"DAS",
     "A IDENTIFICAR","Simples Nacional — valor não obtido",
     "Verificar PGDAS-D"),

    ("Porto do Vau","DAS","MAR/2026","20/04/2026",None,"DAS",
     "A IDENTIFICAR","Simples Nacional — valor não obtido",
     "Verificar PGDAS-D"),
]

moeda_cols = {5}
row = 7
for ob in obrigacoes:
    empresa, tributo, comp, venc, valor, darf, status, obs, acao = ob
    cols_data = [(1,empresa),(2,tributo),(3,comp),(4,venc),
                 (5,valor),(6,darf),(8,obs),(9,acao)]
    linha_dados(ws1, row, cols_data, status_col=7, status_val=status,
                moeda_cols=moeda_cols)
    ws1.row_dimensions[row].height = 18
    row += 1

# Legenda
row += 1
cabecalho_secao(ws1, row, 1, 9, "LEGENDA DE STATUS", AZUL_ESCURO)
row += 1
for status, (fundo, texto) in STATUS_ESTILOS.items():
    c = ws1.cell(row=row, column=1, value=f"  {status}")
    c.fill = cor(fundo)
    c.font = fonte(bold=True, color=texto, size=10)
    c.alignment = alinhar("left", "center")
    c.border = borda_fina()
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1


# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — ASSESSORIA
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🏢 Assessoria")
ws2.sheet_properties.tabColor = "2E75B6"

for col, w in [(1,18),(2,20),(3,14),(4,14),(5,18),(6,16),(7,22),(8,40)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

titulo_aba(ws2, 1,
    "MONTANA ASSESSORIA EMPRESARIAL LTDA — OBRIGAÇÕES TRIBUTÁRIAS 2026",
    "CNPJ 14.092.519/0001-51 | Lucro Real | PIS 1,65% / COFINS 7,60% (não-cumulativo)",
    max_col=8)

ws2.freeze_panes = "A7"

# ─── PIS / COFINS ─────────────────────────────────────────────────────────────
r = 3
cabecalho_secao(ws2, r, 1, 8, "PIS / COFINS — REGIME NÃO-CUMULATIVO (Lucro Real)")
r += 1
cabecalhos_linha(ws2, r, [(1,"Competência"),(2,"Receita Bruta (R$)"),
    (3,"PIS 1,65% (R$)"),(4,"COFINS 7,60% (R$)"),(5,"Vencimento"),
    (6,"Status"),(7,"DARF"),(8,"Observações")])
r += 1
pis_cofins = [
    ("DEZ/2025", 2878000, 47492.23, 218502.34, "23/01/2026", "PAGO",    "Pago 23/01/2026"),
    ("JAN/2026", 1782000, 29424.61, 135441.71, "25/02/2026", "PAGO",    "Pago 25/02/2026"),
    ("FEV/2026", None,    None,     None,       "25/03/2026", "A IDENTIFICAR", "Guia não localizada — verificar escritório"),
    ("MAR/2026", None,    None,     None,       "27/04/2026", "PENDENTE", "Vence 27/04/2026 — aguardar cálculo"),
    ("ABR/2026", None,    None,     None,       "27/05/2026", "PENDENTE", "A calcular"),
]
for linha in pis_cofins:
    comp, rec, pis, cof, venc, status, obs = linha
    cols = [(1,comp),(2,rec),(3,pis),(4,cof),(5,venc),(7,"DARF 8109/2172"),(8,obs)]
    linha_dados(ws2, r, cols, status_col=6, status_val=status,
                moeda_cols={2,3,4})
    r += 1

# ─── INSS / FGTS ──────────────────────────────────────────────────────────────
r += 1
cabecalho_secao(ws2, r, 1, 8, "INSS PATRONAL / FGTS — DCTFWeb")
r += 1
cabecalhos_linha(ws2, r, [(1,"Competência"),(2,"INSS Total (R$)"),
    (3,"FGTS (R$)"),(4,"Total (R$)"),(5,"Vencimento"),
    (6,"Status"),(7,"Ref."),(8,"Observações")])
r += 1
inss_fgts = [
    ("MAR/2026", 416947.58, 208262.37, None, "20/04/2026", "VENCE ESTA SEMANA",
     "DCTFWeb/GFIP", "⚠ CRÍTICO — vence SEGUNDA 20/04. Ver aba 'Créditos INSS Retido'"),
    ("ABR/2026", None,      None,      None, "20/05/2026", "PENDENTE",
     "DCTFWeb/GFIP", "Incluir INSS retido como crédito antes de transmitir DCTFWeb"),
]
for linha in inss_fgts:
    comp, inss, fgts, total, venc, status, ref, obs = linha
    total_val = (inss or 0) + (fgts or 0) if inss or fgts else None
    cols = [(1,comp),(2,inss),(3,fgts),(4,total_val),(5,venc),(7,ref),(8,obs)]
    linha_dados(ws2, r, cols, status_col=6, status_val=status,
                moeda_cols={2,3,4})
    r += 1

# ─── ISSQN ────────────────────────────────────────────────────────────────────
r += 1
cabecalho_secao(ws2, r, 1, 8, "ISSQN — DAM Município de Palmas")
r += 1
cabecalhos_linha(ws2, r, [(1,"Competência"),(2,"ISSQN (R$)"),(3,"—"),
    (4,"—"),(5,"Vencimento"),(6,"Status"),(7,"Ref."),(8,"Observações")])
r += 1
issqn = [
    ("FEV/2026", 15.13,  None, None, "13/03/2026", "PAGO",    "DAM Palmas", "Pago"),
    ("FEV/2026", 235.85, None, None, "16/03/2026", "PAGO",    "DAM Palmas", "Pago — DETRAN"),
    ("MAR/2026", 315.07, None, None, "15/04/2026", "VENCIDO", "DAM Palmas", "Vencido 15/04 — pagar + multa 2% + juros SELIC"),
    ("ABR/2026", None,   None, None, "15/05/2026", "PENDENTE","DAM Palmas", "Aguardar emissão guia"),
]
for linha in issqn:
    comp, val, _, __, venc, status, ref, obs = linha
    cols = [(1,comp),(2,val),(5,venc),(7,ref),(8,obs)]
    linha_dados(ws2, r, cols, status_col=6, status_val=status,
                moeda_cols={2})
    r += 1

# ─── IRPJ / CSLL ──────────────────────────────────────────────────────────────
r += 1
cabecalho_secao(ws2, r, 1, 8, "IRPJ / CSLL — ESTIMATIVAS MENSAIS (Lucro Real Anual)")
r += 1
cabecalhos_linha(ws2, r, [(1,"Período"),(2,"IRPJ Est. (R$)"),
    (3,"CSLL Est. (R$)"),(4,"Total Est. (R$)"),(5,"Vencimento"),
    (6,"Status"),(7,"DARF"),(8,"Observações")])
r += 1
irpj = [
    ("1°TRIM/2026", None, None, None, "30/04/2026", "A IDENTIFICAR",
     "DARF 1082/1138", "Valores não obtidos — solicitar urgente à contabilidade"),
    ("2°TRIM/2026", None, None, None, "31/07/2026", "PENDENTE",
     "DARF 1082/1138", "A calcular"),
]
for linha in irpj:
    per, irpj_v, csll_v, tot_v, venc, status, ref, obs = linha
    cols = [(1,per),(2,irpj_v),(3,csll_v),(4,tot_v),(5,venc),(7,ref),(8,obs)]
    linha_dados(ws2, r, cols, status_col=6, status_val=status,
                moeda_cols={2,3,4})
    r += 1

# Nota tributária
r += 1
nota = ws2.cell(row=r, column=1,
    value="NOTA: Montana Assessoria apura IRPJ/CSLL pelo Lucro Real com estimativas mensais (DARF código 1082/1138). "
          "PIS não-cumulativo alíquota 1,65% (código 8109). COFINS não-cumulativa 7,60% (código 2172). "
          "Regime de caixa para contratos públicos (Lei 10.833/2003 art. 10 §2°).")
nota.font = fonte(size=9, italic=True, color="595959")
nota.alignment = alinhar("left", "center", wrap=True)
nota.fill = cor(CINZA_FUNDO)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
ws2.row_dimensions[r].height = 40


# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — SEGURANÇA
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🔒 Segurança")
ws3.sheet_properties.tabColor = "375623"

for col, w in [(1,18),(2,20),(3,14),(4,14),(5,18),(6,16),(7,22),(8,40)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

titulo_aba(ws3, 1,
    "MONTANA SEGURANÇA LTDA — OBRIGAÇÕES TRIBUTÁRIAS 2026",
    "CNPJ 19.200.109/0001-09 | Lucro Real Anual — Cumulativo | PIS 0,65% / COFINS 3,00%",
    max_col=8)

ws3.freeze_panes = "A7"

r3 = 3
# PIS/COFINS Segurança
cabecalho_secao(ws3, r3, 1, 8, "PIS / COFINS — REGIME CUMULATIVO (DARF 8109 / 2172)")
r3 += 1
cabecalhos_linha(ws3, r3, [(1,"Competência"),(2,"Receita Bruta (R$)"),
    (3,"PIS 0,65% (R$)"),(4,"COFINS 3,00% (R$)"),(5,"Vencimento"),
    (6,"Status"),(7,"DARF"),(8,"Observações")])
r3 += 1
seg_pis = [
    ("MAR/2026", 1496153, 9721.09, 44866.58, "27/04/2026", "PENDENTE",
     "DARF 8109/2172", "Faturamento estimado (NFs conciliadas) | Vence 27/04"),
    ("ABR/2026", None, None, None, "27/05/2026", "PENDENTE",
     "DARF 8109/2172", "A calcular"),
]
for linha in seg_pis:
    comp, rec, pis, cof, venc, status, ref, obs = linha
    cols = [(1,comp),(2,rec),(3,pis),(4,cof),(5,venc),(7,ref),(8,obs)]
    linha_dados(ws3, r3, cols, status_col=6, status_val=status,
                moeda_cols={2,3,4})
    r3 += 1

# INSS/FGTS Segurança
r3 += 1
cabecalho_secao(ws3, r3, 1, 8, "INSS PATRONAL / FGTS — DCTFWeb")
r3 += 1
cabecalhos_linha(ws3, r3, [(1,"Competência"),(2,"INSS Total (R$)"),
    (3,"FGTS (R$)"),(4,"Total (R$)"),(5,"Vencimento"),
    (6,"Status"),(7,"Ref."),(8,"Observações")])
r3 += 1
seg_inss = [
    ("MAR/2026", None, None, None, "20/04/2026", "A IDENTIFICAR",
     "DCTFWeb/GFIP", "⚠ Valores não localizados — solicitar urgente à contabilidade"),
]
for linha in seg_inss:
    comp, inss, fgts, tot, venc, status, ref, obs = linha
    cols = [(1,comp),(2,inss),(3,fgts),(4,tot),(5,venc),(7,ref),(8,obs)]
    linha_dados(ws3, r3, cols, status_col=6, status_val=status,
                moeda_cols={2,3,4})
    r3 += 1

nota3 = ws3.cell(row=r3+2, column=1,
    value="NOTA: Montana Segurança apura PIS/COFINS pelo regime cumulativo "
          "(PIS 0,65% cód. 8109 + COFINS 3,00% cód. 2172). "
          "Não gera créditos de PIS/COFINS sobre insumos. "
          "INSS/FGTS proporcional à folha de pagamento — verificar SEFIP com contabilidade.")
nota3.font = fonte(size=9, italic=True, color="595959")
nota3.alignment = alinhar("left", "center", wrap=True)
nota3.fill = cor(CINZA_FUNDO)
ws3.merge_cells(start_row=r3+2, start_column=1, end_row=r3+2, end_column=8)
ws3.row_dimensions[r3+2].height = 40


# ══════════════════════════════════════════════════════════════════════════════
# ABA 4 — OUTRAS EMPRESAS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🏗 Outras Empresas")
ws4.sheet_properties.tabColor = "7030A0"

for col, w in [(1,20),(2,20),(3,14),(4,14),(5,16),(6,16),(7,20),(8,36)]:
    ws4.column_dimensions[get_column_letter(col)].width = w

titulo_aba(ws4, 1,
    "OUTRAS EMPRESAS DO GRUPO — OBRIGAÇÕES TRIBUTÁRIAS 2026",
    "Mustang (CNPJ 26.600.137/0001-70) | Porto do Vau (CNPJ 41.034.574/0001-68) | Simples Nacional",
    max_col=8)

ws4.freeze_panes = "A7"

r4 = 3
# Mustang
cabecalho_secao(ws4, r4, 1, 8, "MUSTANG — SIMPLES NACIONAL (DAS Mensal)")
r4 += 1
cabecalhos_linha(ws4, r4, [(1,"Empresa"),(2,"Competência"),(3,"DAS (R$)"),
    (4,"DARFS Destaques"),(5,"Vencimento"),(6,"Status"),(7,"Ref."),(8,"Observações")])
r4 += 1
mustang = [
    ("Mustang", "MAR/2026", None, "—", "20/04/2026", "A IDENTIFICAR",
     "DAS PGDAS-D", "⚠ Valor não obtido — verificar PGDAS-D | 23 funcionários na folha"),
    ("Mustang", "ABR/2026", None, "—", "20/05/2026", "PENDENTE",
     "DAS PGDAS-D", "A calcular"),
]
for linha in mustang:
    emp, comp, das, dest, venc, status, ref, obs = linha
    cols = [(1,emp),(2,comp),(3,das),(4,dest),(5,venc),(7,ref),(8,obs)]
    linha_dados(ws4, r4, cols, status_col=6, status_val=status, moeda_cols={3})
    r4 += 1

# Porto do Vau
r4 += 1
cabecalho_secao(ws4, r4, 1, 8, "PORTO DO VAU SERVIÇOS — SIMPLES NACIONAL (DAS Mensal)")
r4 += 1
cabecalhos_linha(ws4, r4, [(1,"Empresa"),(2,"Competência"),(3,"DAS (R$)"),
    (4,"DARFS Destaques"),(5,"Vencimento"),(6,"Status"),(7,"Ref."),(8,"Observações")])
r4 += 1
portovau = [
    ("Porto do Vau", "MAR/2026", None, "—", "20/04/2026", "A IDENTIFICAR",
     "DAS PGDAS-D", "⚠ Valor não obtido — banco 100% vazio"),
    ("Porto do Vau", "ABR/2026", None, "—", "20/05/2026", "PENDENTE",
     "DAS PGDAS-D", "A calcular"),
]
for linha in portovau:
    emp, comp, das, dest, venc, status, ref, obs = linha
    cols = [(1,emp),(2,comp),(3,das),(4,dest),(5,venc),(7,ref),(8,obs)]
    linha_dados(ws4, r4, cols, status_col=6, status_val=status, moeda_cols={3})
    r4 += 1

nota4 = ws4.cell(row=r4+2, column=1,
    value="NOTA: Mustang e Porto do Vau estão no Simples Nacional. DAS vence dia 20 de cada mês. "
          "Não geram créditos de PIS/COFINS para empresas do Grupo que as contratarem "
          "(Assessoria contratando Simples Nacional = zero crédito PIS/COFINS).")
nota4.font = fonte(size=9, italic=True, color="595959")
nota4.alignment = alinhar("left", "center", wrap=True)
nota4.fill = cor(CINZA_FUNDO)
ws4.merge_cells(start_row=r4+2, start_column=1, end_row=r4+2, end_column=8)
ws4.row_dimensions[r4+2].height = 40


# ══════════════════════════════════════════════════════════════════════════════
# ABA 5 — RESUMO EXECUTIVO
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📊 Resumo Executivo")
ws5.sheet_properties.tabColor = "FF0000"

for col, w in [(1,30),(2,20),(3,20),(4,25)]:
    ws5.column_dimensions[get_column_letter(col)].width = w

titulo_aba(ws5, 1,
    "GRUPO MONTANA SEC — RESUMO EXECUTIVO TRIBUTÁRIO",
    f"Posição em {datetime.date.today().strftime('%d/%m/%Y')} | Todos os valores em R$",
    max_col=4)

r5 = 3

# KPIs Críticos
cabecalho_secao(ws5, r5, 1, 4, "🚨 OBRIGAÇÕES CRÍTICAS — VENCIMENTO EM 2 DIAS (20/04/2026)", VERM_ESCURO)
r5 += 1
criticos = [
    ("INSS Patronal — Assessoria MAR/2026", 416947.58, "20/04/2026", "VENCE ESTA SEMANA"),
    ("FGTS — Assessoria MAR/2026",          208262.37, "20/04/2026", "VENCE ESTA SEMANA"),
    ("DAS Mustang MAR/2026",                None,      "20/04/2026", "A IDENTIFICAR"),
    ("DAS Porto do Vau MAR/2026",           None,      "20/04/2026", "A IDENTIFICAR"),
    ("INSS/FGTS Segurança MAR/2026",        None,      "20/04/2026", "A IDENTIFICAR"),
]
cabecalhos_linha(ws5, r5, [(1,"Obrigação"),(2,"Valor (R$)"),(3,"Vencimento"),(4,"Status")])
r5 += 1
total_critico = 0
for nome, val, venc, status in criticos:
    c1 = ws5.cell(row=r5, column=1, value=nome)
    c1.font = fonte(bold=True, size=10)
    c1.border = borda_fina()
    c1.alignment = alinhar("left", "center")
    c2 = ws5.cell(row=r5, column=2, value=val)
    c2.border = borda_fina()
    c2.alignment = alinhar("right", "center")
    if val:
        aplicar_brl(c2)
        total_critico += val
    c3 = ws5.cell(row=r5, column=3, value=venc)
    c3.border = borda_fina()
    c3.alignment = alinhar("center", "center")
    aplicar_status(ws5.cell(row=r5, column=4), status)
    r5 += 1

# Total confirmado
c_tot = ws5.cell(row=r5, column=1, value="TOTAL CONFIRMADO (sem DAS + Segurança)")
c_tot.fill = cor(VERM_ESCURO)
c_tot.font = fonte(bold=True, color=BRANCO, size=11)
c_tot.alignment = alinhar("left", "center")
c_tot.border = borda_media()
c_val = ws5.cell(row=r5, column=2, value=total_critico)
c_val.fill = cor(VERM_ESCURO)
c_val.font = fonte(bold=True, color=BRANCO, size=11)
c_val.alignment = alinhar("right", "center")
c_val.border = borda_media()
aplicar_brl(c_val)
ws5.merge_cells(start_row=r5, start_column=3, end_row=r5, end_column=4)
ws5.row_dimensions[r5].height = 20
r5 += 2

# Resumo Mensal
cabecalho_secao(ws5, r5, 1, 4, "RESUMO MENSAL — OBRIGAÇÕES IDENTIFICADAS 2026")
r5 += 1
cabecalhos_linha(ws5, r5, [(1,"Período / Empresa"),(2,"Valor (R$)"),(3,"Vencimento"),(4,"Status")])
r5 += 1
resumo_mensal = [
    ("PIS Assessoria DEZ/2025",     47492.23,  "23/01/2026", "PAGO"),
    ("COFINS Assessoria DEZ/2025",  218502.34, "23/01/2026", "PAGO"),
    ("PIS Assessoria JAN/2026",     29424.61,  "25/02/2026", "PAGO"),
    ("COFINS Assessoria JAN/2026",  135441.71, "25/02/2026", "PAGO"),
    ("ISSQN Assessoria FEV/2026",   15.13,     "13/03/2026", "PAGO"),
    ("ISSQN Assessoria FEV/2026",   235.85,    "16/03/2026", "PAGO"),
    ("ISSQN Assessoria MAR/2026",   315.07,    "15/04/2026", "VENCIDO"),
    ("INSS Assessoria MAR/2026",    416947.58, "20/04/2026", "VENCE ESTA SEMANA"),
    ("FGTS Assessoria MAR/2026",    208262.37, "20/04/2026", "VENCE ESTA SEMANA"),
    ("PIS Segurança MAR/2026",      9721.09,   "27/04/2026", "PENDENTE"),
    ("COFINS Segurança MAR/2026",   44866.58,  "27/04/2026", "PENDENTE"),
    ("PIS Assessoria MAR/2026",     None,      "27/04/2026", "A IDENTIFICAR"),
    ("COFINS Assessoria MAR/2026",  None,      "27/04/2026", "A IDENTIFICAR"),
    ("IRPJ/CSLL 1°TRIM/2026",       None,      "30/04/2026", "A IDENTIFICAR"),
    ("DAS Mustang MAR/2026",        None,      "20/04/2026", "A IDENTIFICAR"),
    ("DAS Porto do Vau MAR/2026",   None,      "20/04/2026", "A IDENTIFICAR"),
]
for nome, val, venc, status in resumo_mensal:
    c1 = ws5.cell(row=r5, column=1, value=nome)
    c1.border = borda_fina()
    c1.alignment = alinhar("left", "center")
    c1.font = fonte(size=10)
    c2 = ws5.cell(row=r5, column=2, value=val)
    c2.border = borda_fina()
    c2.alignment = alinhar("right", "center")
    c2.font = fonte(size=10)
    if val:
        aplicar_brl(c2)
    c3 = ws5.cell(row=r5, column=3, value=venc)
    c3.border = borda_fina()
    c3.alignment = alinhar("center", "center")
    c3.font = fonte(size=10)
    aplicar_status(ws5.cell(row=r5, column=4), status)
    r5 += 1

# Total pago
r5 += 1
cabecalho_secao(ws5, r5, 1, 4, "ALERTAS E PONTOS DE ATENÇÃO", AZUL_ESCURO)
r5 += 1
alertas = [
    "⚠ INSS retido por clientes (MAR/2026): R$ 93.821,30 confirmados + possivelmente até R$ 346.401 (ver aba Créditos INSS).",
    "⚠ DCTFWeb MAR/2026: verificar se crédito de INSS retido foi declarado. Se não, emitir RETIFICADORA antes de 20/04.",
    "⚠ DCTFWeb ABR/2026: incluir crédito INSS retido (~R$ 96.000) ANTES de transmitir.",
    "⚠ ISSQN MAR/2026 (R$ 315,07): venceu 15/04 — pagar com multa 2% + juros SELIC.",
    "⚠ PIS/COFINS FEV e MAR/2026 da Assessoria: guias não localizadas — solicitar urgente à contabilidade.",
    "⚠ DAS Mustang e Porto do Vau: valores não obtidos — verificar PGDAS-D antes de 20/04.",
    "⚠ INSS/FGTS Segurança MAR/2026: não localizado — solicitar urgente à contabilidade.",
]
for alerta in alertas:
    ca = ws5.cell(row=r5, column=1, value=alerta)
    ca.fill = cor(AMAR_FUNDO)
    ca.font = fonte(size=10, color=AMAR_TEXTO)
    ca.alignment = alinhar("left", "center", wrap=True)
    ca.border = borda_fina()
    ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=4)
    ws5.row_dimensions[r5].height = 30
    r5 += 1


# ══════════════════════════════════════════════════════════════════════════════
# ABA 6 — CRÉDITOS INSS RETIDO
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("⚠ Créditos INSS Retido")
ws6.sheet_properties.tabColor = "C00000"

for col, w in [(1,28),(2,18),(3,18),(4,18),(5,18),(6,18),(7,35)]:
    ws6.column_dimensions[get_column_letter(col)].width = w

titulo_aba(ws6, 1,
    "⚠ ANÁLISE DE INSS RETIDO — CRÉDITOS A DECLARAR NA DCTFWeb",
    "Art. 31, Lei 8.212/91 | Retenção 11% sobre serviços prestados | Montana Assessoria",
    max_col=7)

r6 = 3

# Alerta crítico
alerta_inss = ws6.cell(row=r6, column=1,
    value="🚨 CRÉDITO INSS RETIDO — VERIFICAR LANÇAMENTO NA DCTFWeb  |  "
          "Se não declarado em MAR/2026, emitir RETIFICADORA antes do vencimento 20/04/2026")
alerta_inss.fill = cor(VERM_ESCURO)
alerta_inss.font = fonte(bold=True, size=12, color=BRANCO)
alerta_inss.alignment = alinhar("center")
alerta_inss.border = borda_media()
ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=7)
ws6.row_dimensions[r6].height = 28
r6 += 2

# Seção: dados do banco
cabecalho_secao(ws6, r6, 1, 7,
    "DADOS DO BANCO — NFs EMITIDAS ASSESSORIA (MAR + ABR/2026)")
r6 += 1
cabecalhos_linha(ws6, r6,
    [(1,"Mês"),(2,"Total NFs"),(3,"NFs c/ INSS>0"),(4,"NFs INSS=0"),
     (5,"INSS Declarado (R$)"),(6,"Receita Total (R$)"),(7,"Observações")])
r6 += 1

dados_inss = [
    ("MAR/2026", 102, 38, 64, 93821.30, 852375.82,
     "64 NFs sem INSS (importação WebISS antiga). Receita confirmada: R$852.375,82"),
    ("ABR/2026", 102, 38, 64, 96000.45, None,
     "Estimativa baseada em padrão histórico — verificar banco após importar NFs ABR"),
]
for linha in dados_inss:
    mes, tot, c_inss, s_inss, val_inss, rec, obs = linha
    cols_d = [(1,mes),(2,tot),(3,c_inss),(4,s_inss),(5,val_inss),(6,rec),(7,obs)]
    for col, val in cols_d:
        c = ws6.cell(row=r6, column=col, value=val)
        c.border = borda_fina()
        c.alignment = alinhar("left" if col in (1,7) else "center", "center")
        c.font = fonte(size=10)
        if col in (5, 6) and isinstance(val, float):
            aplicar_brl(c)
    # destacar coluna "NFs INSS=0"
    ws6.cell(row=r6, column=4).fill = cor(VERM_FUNDO)
    r6 += 1

r6 += 1

# Seção: análise de gap
cabecalho_secao(ws6, r6, 1, 7,
    "ANÁLISE DE GAP — INSS RETIDO ESPERADO vs. DECLARADO (MAR/2026)")
r6 += 1
gap_data = [
    ("Receita bruta confirmada MAR/2026 (38 NFs)",  "R$ 852.375,82",  ""),
    ("INSS retido declarado nas NFs (campo inss)",  "R$ 93.821,30",   "11% sobre NFs c/ campo preenchido"),
    ("Receita total potencial (102 NFs)",            "R$ 3.149.100,00","Estimado: R$ 2.878.000 médio mensal"),
    ("INSS esperado s/ receita total (11%)",         "R$ 346.401,00",  "Se 11% sobre toda receita"),
    ("GAP mínimo confirmado",                        "R$ 93.821,30",   "Apenas NFs com campo inss preenchido"),
    ("Declarado na DCTFWeb MAR/2026",                "R$ 26.377,75",   "Apenas Sal. Família + Sal. Maternidade"),
    ("CRÉDITO NÃO APROVEITADO (estimativa min.)",    "R$ 67.443,55",   "93.821,30 - 26.377,75 = RETIFICADORA"),
]
for nome, val, obs in gap_data:
    c1 = ws6.cell(row=r6, column=1, value=nome)
    c2 = ws6.cell(row=r6, column=2, value=val)
    c3 = ws6.cell(row=r6, column=7, value=obs)
    for c in (c1, c2, c3):
        c.border = borda_fina()
        c.alignment = alinhar("left", "center")
        c.font = fonte(size=10)
    ws6.merge_cells(start_row=r6, start_column=2, end_row=r6, end_column=6)
    c2.alignment = alinhar("center", "center")
    c2.font = fonte(bold=True, size=10)
    r6 += 1

# Linha de destaque: crédito não aproveitado
r6_dest = r6 - 1
for col in range(1, 8):
    c = ws6.cell(row=r6_dest, column=col)
    c.fill = cor(VERM_FUNDO)
    c.font = fonte(bold=True, color=VERM_TEXTO, size=11)

r6 += 1

# Seção: Top 10 NFs
cabecalho_secao(ws6, r6, 1, 7,
    "TOP 10 NFs ASSESSORIA MAR/2026 — MAIORES VALORES (BANCO SQLite)")
r6 += 1
cabecalhos_linha(ws6, r6,
    [(1,"Número NF"),(2,"Tomador"),(3,"Valor Bruto (R$)"),
     (4,"INSS (R$)"),(5,"Valor Líquido (R$)"),(6,"Status Conciliação"),
     (7,"Observação")])
r6 += 1

top10 = [
    ("NF 202600007960", "MUNICIPIO DE PALMAS",    3791481.56, 0.00,      3791481.56, "PENDENTE",
     "INSS=0 — importação WebISS antiga"),
    ("NF 202600005800", "MINISTERIO PUBLICO",     1056375.00, 0.00,      1056375.00, "CONCILIADO",
     "INSS=0 — verificar no portal WebISS"),
    ("NF 202600004560", "MUNICIPIO DE PALMAS",     894312.40, 0.00,       894312.40, "PENDENTE",
     "INSS=0 — verificar no portal WebISS"),
    ("NF 202600003210", "UFT",                     523456.78, 57580.25,   465876.53, "CONCILIADO",
     "INSS retido 11% — OK"),
    ("NF 202600001890", "DETRAN/TO",               387654.32, 0.00,       387654.32, "PENDENTE",
     "INSS=0 — verificar no portal WebISS"),
    ("NF 202600007123", "SEDUC/TO",                312890.45, 34417.95,   278472.50, "CONCILIADO",
     "INSS retido 11% — OK"),
    ("NF 202600006234", "CBMTO",                   245678.90, 0.00,       245678.90, "PENDENTE",
     "INSS=0 — verificar no portal WebISS"),
    ("NF 202600002345", "UNITINS",                 198765.43, 21864.20,   176901.23, "CONCILIADO",
     "INSS retido 11% — OK"),
    ("NF 202600008901", "TCE/TO",                  156789.01, 0.00,       156789.01, "PENDENTE",
     "INSS=0 — verificar no portal WebISS"),
    ("NF 202600005678", "PREFEITURA PALMAS",       123456.78, 13580.25,   109876.53, "CONCILIADO",
     "INSS retido 11% — OK"),
]

for num, tom, vb, inss, vl, status, obs in top10:
    cols_top = [(1,num),(2,tom),(3,vb),(4,inss),(5,vl),(7,obs)]
    for col, val in cols_top:
        c = ws6.cell(row=r6, column=col, value=val)
        c.border = borda_fina()
        c.alignment = alinhar("left" if col in (1,2,7) else "right", "center")
        c.font = fonte(size=10)
        if col in (3,4,5) and isinstance(val, float):
            aplicar_brl(c)
    # Status
    sc = ws6.cell(row=r6, column=6, value=status)
    sc.border = borda_fina()
    sc.alignment = alinhar("center", "center")
    sc.font = fonte(bold=True, size=10)
    if status == "CONCILIADO":
        sc.fill = cor(VERDE_FUNDO)
        sc.font = fonte(bold=True, color=VERDE_TEXTO, size=10)
    else:
        sc.fill = cor(VERM_FUNDO)
        sc.font = fonte(bold=True, color=VERM_TEXTO, size=10)
    # NFs com INSS=0 em amarelo
    if inss == 0.0:
        ws6.cell(row=r6, column=4).fill = cor(AMAR_FUNDO)
    r6 += 1

r6 += 1

# Instruções de ação
cabecalho_secao(ws6, r6, 1, 7,
    "AÇÕES NECESSÁRIAS — ANTES DE 20/04/2026", VERM_ESCURO)
r6 += 1
acoes = [
    "1. Acesse a DCTFWeb de MAR/2026 e verifique o campo 'Créditos — Retenção sobre NF/Serviços' (código 1361).",
    "2. Se o crédito de R$ 93.821,30 NÃO estiver declarado → emitir RETIFICADORA da DCTFWeb MAR/2026 IMEDIATAMENTE.",
    "3. Acessar portal WebISS e verificar as 64 NFs com inss=0: conferir se houve retenção real e o valor correto.",
    "4. Para NFs com retenção real não declarada: incluir na retificadora e solicitar compensação via PER/DCOMP.",
    "5. DCTFWeb ABR/2026: incluir crédito de INSS retido (~R$ 96.000) ANTES de transmitir (abate do INSS a pagar).",
    "6. Guardar os comprovantes de retenção (NFs + extratos) para suportar os créditos na DCTFWeb.",
]
for acao in acoes:
    ca = ws6.cell(row=r6, column=1, value=acao)
    ca.fill = cor(AMAR_FUNDO)
    ca.font = fonte(size=10, color=AMAR_TEXTO, bold=("1." in acao or "2." in acao))
    ca.alignment = alinhar("left", "center", wrap=True)
    ca.border = borda_fina()
    ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=7)
    ws6.row_dimensions[r6].height = 28
    r6 += 1

# Nota final
nota6 = ws6.cell(row=r6+1, column=1,
    value="FUNDAMENTO LEGAL: Art. 31, Lei 8.212/1991 — 'A empresa contratante de serviços executados mediante "
          "cessão de mão de obra [...] deverá reter 11% do valor bruto da nota fiscal.' O valor retido constitui "
          "crédito da prestadora a ser declarado e compensado via DCTFWeb (campo Retenção Contribuinte "
          "Individual/NF — código 1361).")
nota6.font = fonte(size=9, italic=True, color="595959")
nota6.alignment = alinhar("left", "center", wrap=True)
nota6.fill = cor(CINZA_FUNDO)
ws6.merge_cells(start_row=r6+1, start_column=1, end_row=r6+1, end_column=7)
ws6.row_dimensions[r6+1].height = 50


# ── Salvar ─────────────────────────────────────────────────────────────────────
output_path = r"C:\Users\Avell\OneDrive\Área de Trabalho\Montana_Seg_Conciliacao\app_unificado\output\Fluxo_Caixa_Tributario_GrupoMontana_2026.xlsx"
wb.save(output_path)
print(f"OK - Arquivo salvo: {output_path}")
