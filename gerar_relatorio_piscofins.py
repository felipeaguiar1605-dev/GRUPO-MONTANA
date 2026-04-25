#!/usr/bin/env python3
"""Gera relatório final PIS/COFINS Assessoria Mar/2026"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT = '/sessions/confident-lucid-rubin/mnt/app_unificado/Apuracao_PISCOFINS_Assessoria_202603_FINAL.xlsx'

# ── Paleta ──────────────────────────────────────────────────────────────────
AZUL_ESC  = '1F3864'
AZUL_MED  = '2E75B6'
AZUL_CLA  = 'D6E4F0'
VERDE_ESC = '1B5E20'
VERDE_CLA = 'E8F5E9'
AMBAR_CLA = 'FFF8E1'
CINZA_CLA = 'F5F5F5'
BRANCO    = 'FFFFFF'
VERMELHO  = 'C62828'
LARANJA   = 'F57C00'

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def font(bold=False, color='000000', size=11, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)

def border_thin():
    s = Side(style='thin', color='BDBDBD')
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium_bottom():
    return Border(bottom=Side(style='medium', color='1F3864'))

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def right():
    return Alignment(horizontal='right', vertical='center')

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

FMT_BRL  = '#,##0.00'
FMT_PCT  = '0.00%'
FMT_DATA = 'DD/MM/AAAA'

# ── Dados ────────────────────────────────────────────────────────────────────
# 93 créditos originais da planilha v11 (aba Tributável)
TRIBUTAVEIS_ORIG = [
    ('18/03/2026', 430496.43, '25053083000108', 'SECRETARIA DA EDUCACAO',                         'Contrato SEDUC Limpeza/Copeiragem'),
    ('05/03/2026', 235190.76, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('18/03/2026', 193359.33, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('18/03/2026', 193359.33, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('09/03/2026', 185267.18, '25053083000108', 'SECRETARIA DA EDUCACAO',                         'Contrato SEDUC Limpeza/Copeiragem'),
    ('31/03/2026', 176331.50, '05149726000104', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Resgate Depósito Garantia / NF 246'),
    ('09/03/2026', 163799.68, '25053083000108', 'SECRETARIA DA EDUCACAO',                         'Contrato SEDUC Limpeza/Copeiragem'),
    ('19/03/2026', 147341.41, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('18/03/2026', 143865.79, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('18/03/2026', 137586.05, '25053083000108', 'SECRETARIA DA EDUCACAO',                         'Contrato SEDUC Limpeza/Copeiragem'),
    ('19/03/2026', 122784.51, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('17/03/2026', 104815.76, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('31/03/2026', 102556.43, '25053117000164', 'SESAU-TO',                                        'Resgate Depósito Garantia / NF 191'),
    ('19/03/2026',  99390.38, '25053083000108', 'SECRETARIA DA EDUCACAO',                         'Contrato SEDUC Limpeza/Copeiragem'),
    ('19/03/2026',  73670.71, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('18/03/2026',  62832.49, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('18/03/2026',  62442.18, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('12/03/2026',  56545.49, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('03/03/2026',  51013.11, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('09/03/2026',  50000.00, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('17/03/2026',  49113.81, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('26/03/2026',  49113.81, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('17/03/2026',  49113.80, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('18/03/2026',  48726.00, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('03/03/2026',  47500.00, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('03/03/2026',  42618.71, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('03/03/2026',  41933.74, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('11/03/2026',  39862.12, '05149726000104', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'NF 202600000000295'),
    ('03/03/2026',  39653.99, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('18/03/2026',  37024.25, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('17/03/2026',  36016.79, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('16/03/2026',  35862.16, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('19/03/2026',  35862.16, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('18/03/2026',  35112.08, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('11/03/2026',  30000.00, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('11/03/2026',  28644.51, '01637536000185', 'UNITINS',                                         'OB 2026OB000515'),
    ('06/03/2026',  28474.95, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('06/03/2026',  28474.95, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('16/03/2026',  28474.95, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('19/03/2026',  26194.03, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('16/03/2026',  24556.90, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('19/03/2026',  24556.90, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('26/03/2026',  24556.90, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('11/03/2026',  21491.95, '05149726000104', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'NF 202600000000294'),
    ('09/03/2026',  21467.50, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('11/03/2026',  20902.16, '05149726000104', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'NF 202600000000296'),
    ('03/03/2026',  20749.12, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('16/03/2026',  19645.52, '05278848000109', 'PREVIPALMAS',                                     'Contrato PREVIPALMAS'),
    ('18/03/2026',  19038.27, '25053117000164', 'SESAU-TO',                                        'Contrato SESAU 178/2022'),
    ('18/03/2026',  18704.15, '25053117000164', 'SESAU-TO',                                        'Contrato SESAU 178/2022'),
    ('06/03/2026',  18228.77, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('06/03/2026',  18228.77, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('16/03/2026',  18008.39, '05016202000145', 'SEMARH-TO',                                       'Contrato SEMARH 32/2024'),
    ('31/03/2026',  16736.34, '38178825000173', 'UFNT',                                            'Resgate Depósito Garantia / NF 280'),
    ('11/03/2026',  16428.01, '01637536000185', 'UNITINS',                                         'OB 2026OB000513'),
    ('11/03/2026',  16225.88, '05149726000104', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'NF 202600000000297'),
    ('18/03/2026',  16143.55, '05278848000109', 'PREVIPALMAS',                                     'Contrato PREVIPALMAS'),
    ('18/03/2026',  15637.97, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('16/03/2026',  13097.01, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('16/03/2026',  12859.84, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('26/03/2026',  12859.84, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('26/03/2026',  11697.07, '25053117000164', 'SESAU-TO',                                        'Contrato SESAU 178/2022'),
    ('18/03/2026',  10580.00, '01637536000185', 'UNITINS',                                         'OB 2026OB000573'),
    ('06/03/2026',  10000.00, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('06/03/2026',  10000.00, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('26/03/2026',   9667.72, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('16/03/2026',   9563.24, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('18/03/2026',   8833.31, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('06/03/2026',   8395.00, '01637536000185', 'UNITINS',                                         'OB 2026OB000492'),
    ('06/03/2026',   8312.40, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('06/03/2026',   8312.40, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('16/03/2026',   8185.64, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('16/03/2026',   8185.63, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('18/03/2026',   6961.88, '01637536000185', 'UNITINS',                                         'Contrato UNITINS 003/2023'),
    ('16/03/2026',   6548.50, '05149726000104', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'NF 202600000000252'),
    ('18/03/2026',   6480.76, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('16/03/2026',   6139.23, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('16/03/2026',   6139.22, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('16/03/2026',   6002.80, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('18/03/2026',   5911.37, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('06/03/2026',   5000.00, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('06/03/2026',   5000.00, '01060887000195', 'DETRAN-TO',                                       'Contrato DETRAN 41/2023'),
    ('16/03/2026',   4502.11, '',               'MUNICÍPIO DE PALMAS',                             'OB Prefeitura Palmas'),
    ('16/03/2026',   4502.10, '01637536000185', 'UNITINS',                                         'OB 2026DF800577'),
    ('11/03/2026',   4200.00, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('16/03/2026',   3429.28, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('16/03/2026',   3274.25, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('11/03/2026',   3262.39, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('11/03/2026',   3262.39, '38178825000173', 'UFNT',                                            'Contrato UFNT 30/2022'),
    ('18/03/2026',   3179.36, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('16/03/2026',   2182.84, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('16/03/2026',   2182.83, '14092519000151', 'FUNDACAO UNIV FEDERAL DO TOCANTINS (UFT)',        'Contrato UFT 16/2025'),
    ('17/03/2026',   1637.13, '',               'MUNICÍPIO DE PALMAS',                             'OB Prefeitura Palmas'),
]

# 3 créditos reclassificados de DIFERIDO → TRIBUTÁVEL (confirmados pelo proprietário)
DIFERIDO_RECLASSIFICADO = [
    ('31/03/2026', 430496.43, '01060887000195', 'DETRAN-TO',  'Pagamento ref. Fev/2026 — reclassif. de DIFERIDO'),
    ('05/03/2026', 277281.38, '01637536000185', 'UNITINS',    'TED Gov. Estado TO — reclassif. de DIFERIDO'),
    ('06/03/2026',  17124.67, '05016202000145', 'SEMARH-TO',  'TED Gov. Estado TO — reclassif. de DIFERIDO'),
]

# Resumo NÃO TRIBUTA por categoria
NAO_TRIBUTA_CATS = [
    ('Lançamentos de Saldo (S A L D O)',            2,  1083192.13),
    ('BB Rende Fácil / Aplicações',                12,  1850289.22),
    ('Transferências Internas — Grupo Montana',    10,  1101441.76),
    ('Resgates Depósito Garantia',                  3,   368226.92),
    ('Resgate FI BRB Federal Invest',               1,    47500.00),
    ('Desbloqueio Judicial (BACEN JUD)',             1,    28086.66),
    ('Repasses pessoa física / poupança',           3,      805.00),
    ('CRD Juros CDB Automático',                    3,       24.81),
    ('BB Rende Fácil (INTERNO duplicado)',           1,   214320.55),
    ('Rede Nacional (PIX interno)',                  1,     1222.43),
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_row(ws, row, cols, texts, bg, fg='FFFFFF', sz=11, bold=True):
    for c, txt in zip(cols, texts):
        cell = ws.cell(row=row, column=c, value=txt)
        cell.font = font(bold=bold, color=fg, size=sz)
        cell.fill = fill(bg)
        cell.alignment = center()
        cell.border = border_thin()

def data_cell(ws, row, col, value, fmt=None, bold=False, bg=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold)
    cell.border = border_thin()
    if fmt:
        cell.number_format = fmt
    if bg:
        cell.fill = fill(bg)
    cell.alignment = align or left()
    return cell

def money(ws, row, col, value, bold=False, bg=None):
    data_cell(ws, row, col, value, FMT_BRL, bold=bold, bg=bg, align=right())

def title_block(ws, row, text, subtext=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name='Arial', bold=True, color=BRANCO, size=14)
    c.fill = fill(AZUL_ESC)
    c.alignment = center()
    if subtext:
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=8)
        c2 = ws.cell(row=row+1, column=1, value=subtext)
        c2.font = Font(name='Arial', italic=True, color=AZUL_MED, size=10)
        c2.fill = fill(AZUL_CLA)
        c2.alignment = center()


# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── ABA 1: RESUMO EXECUTIVO ────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Resumo Executivo'
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 22

for c in range(1, 9):
    set_col_width(ws1, c, [3, 28, 20, 18, 18, 18, 18, 10][c-1])

# Título
title_block(ws1, 1,
    'APURAÇÃO PIS/COFINS — MARÇO/2026',
    'Montana Assessoria em Segurança Ltda  ·  Lucro Real Não-Cumulativo  ·  Regime de Caixa  ·  Emitido em ' + datetime.now().strftime('%d/%m/%Y'))

# ── Bloco: Composição dos créditos ──
ws1.row_dimensions[4].height = 20
header_row(ws1, 4, [2,3,4,5], ['COMPOSIÇÃO DOS CRÉDITOS', 'Qtd', 'Valor (R$)', '% Total'],
           AZUL_MED, sz=10)
ws1.merge_cells('B4:B4')

comp_data = [
    ('Créditos TRIBUTÁVEIS (planilha)',         93,  4142079.92,  True,  VERDE_CLA),
    ('+ Reclassificados de DIFERIDO',            3,   724902.48,  True,  VERDE_CLA),
    ('  DETRAN-TO — ref. Fev/2026',             '',  430496.43,  False, BRANCO),
    ('  UNITINS — confirmado',                  '',  277281.38,  False, BRANCO),
    ('  SEMARH-TO — confirmado',                '',   17124.67,  False, BRANCO),
    ('Créditos NÃO TRIBUTÁVEIS',               37,  4846544.50, False, CINZA_CLA),
    ('TOTAL GERAL DO EXTRATO',                 133,  9713526.90, True,  AZUL_CLA),
]
r = 5
for label, qtd, val, bold, bg in comp_data:
    ws1.cell(r, 2, label).font    = font(bold=bold)
    ws1.cell(r, 2).fill           = fill(bg)
    ws1.cell(r, 2).border         = border_thin()
    ws1.cell(r, 2).alignment      = left()
    if qtd != '':
        ws1.cell(r, 3, qtd).number_format = '#,##0'
    ws1.cell(r, 3).fill   = fill(bg); ws1.cell(r,3).border = border_thin(); ws1.cell(r,3).alignment = center()
    ws1.cell(r, 4, val).number_format = FMT_BRL
    ws1.cell(r, 4).fill   = fill(bg); ws1.cell(r,4).border = border_thin(); ws1.cell(r,4).alignment = right()
    ws1.cell(r, 4).font   = font(bold=bold)
    if label == 'TOTAL GERAL DO EXTRATO':
        ws1.cell(r, 5, f'=D{r}/D{r}').number_format = FMT_PCT
    elif val and label not in ('  DETRAN-TO — ref. Fev/2026','  UNITINS — confirmado','  SEMARH-TO — confirmado'):
        ws1.cell(r, 5, f'=D{r}/D{11}').number_format = FMT_PCT
    ws1.cell(r, 5).fill  = fill(bg); ws1.cell(r,5).border = border_thin(); ws1.cell(r,5).alignment = right()
    r += 1

# ── Bloco: BASE E APURAÇÃO ──
r += 1
ws1.row_dimensions[r].height = 20
header_row(ws1, r, [2,3,4,5], ['APURAÇÃO DO IMPOSTO', '', 'Valor (R$)', ''],
           AZUL_ESC, sz=11)
r += 1

apuracao = [
    ('BASE TRIBUTÁVEL TOTAL',                   '',  4866982.40, True,  VERDE_CLA,  True),
    ('  Planilha v11 (93 créditos)',             '',  4142079.92, False, BRANCO,     False),
    ('  + DIFERIDO reclassificado (3 créditos)','',   724902.48, False, BRANCO,     False),
    ('',None,None,False,BRANCO,False),
    ('PIS — débito (1,65%)',                    '',    80305.21, False, CINZA_CLA,  False),
    ('  (−) Crédito PIS de entrada',            '',    -9741.72, False, BRANCO,     False),
    ('PIS A RECOLHER',                          '',    70563.49, True,  AMBAR_CLA,  True),
    ('',None,None,False,BRANCO,False),
    ('COFINS — débito (7,60%)',                 '',   369890.66, False, CINZA_CLA,  False),
    ('  (−) Crédito COFINS de entrada',         '',   -44897.70, False, BRANCO,     False),
    ('COFINS A RECOLHER',                       '',   324992.96, True,  AMBAR_CLA,  True),
    ('',None,None,False,BRANCO,False),
    ('TOTAL PIS + COFINS A RECOLHER',           '',   395556.45, True,  VERDE_ESC,  True),
]
for label, _, val, bold, bg, highlight in apuracao:
    ws1.cell(r, 2, label).font      = font(bold=bold, color=BRANCO if highlight and bg==VERDE_ESC else '000000')
    ws1.cell(r, 2).fill             = fill(bg)
    ws1.cell(r, 2).border           = border_thin()
    ws1.cell(r, 2).alignment        = left()
    if val is not None:
        c = ws1.cell(r, 4, val)
        c.number_format = FMT_BRL
        c.fill          = fill(bg)
        c.border        = border_thin()
        c.alignment     = right()
        c.font          = font(bold=bold, color=BRANCO if highlight and bg==VERDE_ESC else ('C62828' if val < 0 else '000000'))
    ws1.cell(r, 3).fill  = fill(bg); ws1.cell(r,3).border = border_thin()
    ws1.cell(r, 5).fill  = fill(bg); ws1.cell(r,5).border = border_thin()
    r += 1

# ── Bloco: DARF ──
r += 1
ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws1.cell(r, 2, '⚠  DARF — Vencimento: 27/04/2026  (dia 25 cai em sábado → próximo dia útil)')
c.font      = Font(name='Arial', bold=True, color='7B3F00', size=10)
c.fill      = fill('FFF3CD')
c.alignment = center()
for col in [2,3,4,5]:
    ws1.cell(r, col).border = border_thin()

r += 1
darf_rows = [
    ('PIS — código DARF: 6912',   70563.49),
    ('COFINS — código DARF: 5856', 324992.96),
]
for label, val in darf_rows:
    ws1.cell(r, 2, label).font      = font(bold=True)
    ws1.cell(r, 2).fill             = fill(AMBAR_CLA)
    ws1.cell(r, 2).border           = border_thin()
    ws1.cell(r, 2).alignment        = left()
    ws1.cell(r, 4, val).number_format = FMT_BRL
    ws1.cell(r, 4).font             = font(bold=True)
    ws1.cell(r, 4).fill             = fill(AMBAR_CLA)
    ws1.cell(r, 4).border           = border_thin()
    ws1.cell(r, 4).alignment        = right()
    ws1.cell(r, 3).fill             = fill(AMBAR_CLA); ws1.cell(r,3).border = border_thin()
    ws1.cell(r, 5).fill             = fill(AMBAR_CLA); ws1.cell(r,5).border = border_thin()
    r += 1

# rodapé
r += 2
ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
c = ws1.cell(r, 2, f'Fonte: Planilha PIS_COFINS_CAIXA_assessoria_202603_v11.xlsx  ·  Classificações DIFERIDO confirmadas pelo proprietário em 18/04/2026')
c.font      = Font(name='Arial', italic=True, color='757575', size=9)
c.alignment = left()


# ── ABA 2: CRÉDITOS TRIBUTÁVEIS ────────────────────────────────────────────
ws2 = wb.create_sheet('Créditos Tributáveis')
ws2.sheet_view.showGridLines = False

col_widths2 = [3, 6, 14, 16, 20, 40, 35, 10]
for i, w in enumerate(col_widths2, 1):
    set_col_width(ws2, i, w)

ws2.row_dimensions[1].height = 36
ws2.row_dimensions[2].height = 18
title_block(ws2, 1,
    'CRÉDITOS TRIBUTÁVEIS — MARÇO/2026',
    '96 créditos  ·  Base: R$ 4.866.982,40  ·  PIS 1,65% + COFINS 7,60% = 9,25%')

header_row(ws2, 3, [2,3,4,5,6,7,8],
    ['#','Data','Valor (R$)','CNPJ','Tomador / Pagador','Contrato / Referência','Origem'],
    AZUL_MED, sz=10)

todos = [(d, v, cnpj, tom, ref, 'Planilha v11') for d,v,cnpj,tom,ref in TRIBUTAVEIS_ORIG]
todos += [(d, v, cnpj, tom, ref, '★ Reclassif. DIFERIDO') for d,v,cnpj,tom,ref in DIFERIDO_RECLASSIFICADO]
todos.sort(key=lambda x: (x[0][-4:]+x[0][3:5]+x[0][:2]))  # sort by date

total_trib = sum(v for _,v,*_ in todos)
r = 4
for i, (data, val, cnpj, tom, ref, origem) in enumerate(todos, 1):
    bg = CINZA_CLA if i % 2 == 0 else BRANCO
    is_reclassif = origem.startswith('★')
    if is_reclassif:
        bg = VERDE_CLA
    ws2.cell(r, 2, i).fill = fill(bg); ws2.cell(r,2).border = border_thin(); ws2.cell(r,2).alignment = center(); ws2.cell(r,2).font = font()
    ws2.cell(r, 3, data).fill = fill(bg); ws2.cell(r,3).border = border_thin(); ws2.cell(r,3).alignment = center(); ws2.cell(r,3).font = font()
    money(ws2, r, 4, val, bg=bg)
    ws2.cell(r, 5, cnpj).fill = fill(bg); ws2.cell(r,5).border = border_thin(); ws2.cell(r,5).alignment = center(); ws2.cell(r,5).font = font(size=9)
    ws2.cell(r, 6, tom).fill = fill(bg); ws2.cell(r,6).border = border_thin(); ws2.cell(r,6).alignment = left(); ws2.cell(r,6).font = font(bold=is_reclassif)
    ws2.cell(r, 7, ref).fill = fill(bg); ws2.cell(r,7).border = border_thin(); ws2.cell(r,7).alignment = left(); ws2.cell(r,7).font = font(size=9)
    ws2.cell(r, 8, origem).fill = fill(bg); ws2.cell(r,8).border = border_thin(); ws2.cell(r,8).alignment = center(); ws2.cell(r,8).font = font(bold=is_reclassif, color=VERDE_ESC if is_reclassif else '000000', size=9)
    r += 1

# Total
header_row(ws2, r, [2,3,4,5,6,7,8], ['TOTAL','96',total_trib,'','','',''], AZUL_ESC)
ws2.cell(r, 4).number_format = FMT_BRL
ws2.cell(r, 3).alignment = center()
ws2.row_dimensions[r].height = 18

# Legenda
r += 2
ws2.cell(r, 2, '★ Fundo verde = crédito reclassificado de DIFERIDO para TRIBUTÁVEL (confirmado pelo proprietário em 18/04/2026)').font = Font(name='Arial', italic=True, color=VERDE_ESC, size=9)


# ── ABA 3: POR TOMADOR ─────────────────────────────────────────────────────
ws3 = wb.create_sheet('Por Tomador')
ws3.sheet_view.showGridLines = False

col_widths3 = [3, 42, 20, 20, 20, 14, 14]
for i, w in enumerate(col_widths3, 1):
    set_col_width(ws3, i, w)

ws3.row_dimensions[1].height = 36
ws3.row_dimensions[2].height = 18
title_block(ws3, 1,
    'RECEITA TRIBUTÁVEL POR TOMADOR — MARÇO/2026',
    'Consolidado  ·  Lucro Real Não-Cumulativo')

header_row(ws3, 3, [2,3,4,5,6,7],
    ['Tomador / Pagador','CNPJ','Total Recebido (R$)','% da Base','PIS (1,65%)','COFINS (7,60%)'],
    AZUL_MED, sz=10)

# Consolidar por tomador
from collections import defaultdict
por_tom = defaultdict(lambda: [set(), 0.0])
for data, val, cnpj, tom, ref, origem in todos:
    key = (tom, cnpj)
    por_tom[key][0].add(cnpj)
    por_tom[key][1] += val

tomadores = sorted(por_tom.items(), key=lambda x: -x[1][1])
r = 4
for i, ((tom, cnpj), (cnpjs, total)) in enumerate(tomadores, 1):
    bg = CINZA_CLA if i % 2 == 0 else BRANCO
    pct_ref_row = r
    ws3.cell(r, 2, tom).fill = fill(bg); ws3.cell(r,2).border = border_thin(); ws3.cell(r,2).font = font(bold=True); ws3.cell(r,2).alignment = left()
    ws3.cell(r, 3, cnpj).fill = fill(bg); ws3.cell(r,3).border = border_thin(); ws3.cell(r,3).font = font(size=9); ws3.cell(r,3).alignment = center()
    money(ws3, r, 4, total, bg=bg)
    ws3.cell(r, 5, f'=D{r}/D{r+len(tomadores)-i+1+1}').number_format = FMT_PCT; ws3.cell(r,5).fill = fill(bg); ws3.cell(r,5).border = border_thin(); ws3.cell(r,5).alignment = right(); ws3.cell(r,5).font = font()
    ws3.cell(r, 6, f'=D{r}*0.0165').number_format = FMT_BRL; ws3.cell(r,6).fill = fill(bg); ws3.cell(r,6).border = border_thin(); ws3.cell(r,6).alignment = right(); ws3.cell(r,6).font = font()
    ws3.cell(r, 7, f'=D{r}*0.076').number_format  = FMT_BRL; ws3.cell(r,7).fill = fill(bg); ws3.cell(r,7).border = border_thin(); ws3.cell(r,7).alignment = right(); ws3.cell(r,7).font = font()
    r += 1

# Total
header_row(ws3, r, [2,3,4,5,6,7],
    ['TOTAL','',f'=SUM(D4:D{r-1})','100%',f'=SUM(F4:F{r-1})',f'=SUM(G4:G{r-1})'],
    AZUL_ESC)
for col in [4,6,7]:
    ws3.cell(r, col).number_format = FMT_BRL


# ── ABA 4: NÃO TRIBUTA ─────────────────────────────────────────────────────
ws4 = wb.create_sheet('Não Tributa')
ws4.sheet_view.showGridLines = False

col_widths4 = [3, 46, 12, 20, 20]
for i, w in enumerate(col_widths4, 1):
    set_col_width(ws4, i, w)

ws4.row_dimensions[1].height = 36
ws4.row_dimensions[2].height = 18
title_block(ws4, 1,
    'CRÉDITOS NÃO TRIBUTÁVEIS — MARÇO/2026',
    '37 créditos  ·  R$ 4.846.544,50  ·  NÃO integram a base de PIS/COFINS')

header_row(ws4, 3, [2,3,4,5],
    ['Categoria','Qtd','Valor (R$)','Motivo / Enquadramento'],
    AZUL_MED, sz=10)

total_nt = 0
r = 4
for i, (cat, qtd, val) in enumerate(NAO_TRIBUTA_CATS, 1):
    bg = CINZA_CLA if i % 2 == 0 else BRANCO
    ws4.cell(r, 2, cat).fill = fill(bg); ws4.cell(r,2).border = border_thin(); ws4.cell(r,2).font = font(); ws4.cell(r,2).alignment = left()
    ws4.cell(r, 3, qtd).fill = fill(bg); ws4.cell(r,3).border = border_thin(); ws4.cell(r,3).font = font(); ws4.cell(r,3).alignment = center(); ws4.cell(r,3).number_format = '#,##0'
    money(ws4, r, 4, val, bg=bg)
    motivo = {
        'Lançamentos de Saldo (S A L D O)':        'Controle interno de saldo — não representa entrada de recurso',
        'BB Rende Fácil / Aplicações':             'Retorno de aplicação financeira — não é receita de serviço',
        'Transferências Internas — Grupo Montana': 'Movimentação entre empresas do grupo — não é receita',
        'Resgates Depósito Garantia':              'Devolução de caução contratual — não é receita de serviço',
        'Resgate FI BRB Federal Invest':           'Resgate de fundo de investimento — não é receita',
        'Desbloqueio Judicial (BACEN JUD)':        'Desbloqueio judicial de conta — natureza não tributável',
        'Repasses pessoa física / poupança':       'Repasse avulso / poupança — não é receita de prestação de serviço',
        'CRD Juros CDB Automático':               'Rendimento financeiro de CDB — não-cumulativo: não tributa PIS/COFINS',
        'BB Rende Fácil (INTERNO duplicado)':      'Entrada classificada como INTERNO pelo sistema (duplicidade de lançamento)',
        'Rede Nacional (PIX interno)':             'PIX recebido de entidade interna — classificado como INTERNO',
    }.get(cat, '')
    ws4.cell(r, 5, motivo).fill = fill(bg); ws4.cell(r,5).border = border_thin(); ws4.cell(r,5).font = font(size=9, italic=True); ws4.cell(r,5).alignment = left()
    total_nt += val
    r += 1

header_row(ws4, r, [2,3,4,5],
    ['TOTAL NÃO TRIBUTÁVEL', f'=SUM(C4:C{r-1})', f'=SUM(D4:D{r-1})', ''], AZUL_ESC)
ws4.cell(r, 4).number_format = FMT_BRL

wb.save(OUTPUT)
print(f'Salvo: {OUTPUT}')
print(f'Total tributável: R$ {total_trib:,.2f}')
